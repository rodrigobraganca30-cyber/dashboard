import openpyxl
import glob
import os
import json
import unicodedata
from datetime import datetime
from collections import Counter

# --- CONFIGURAÇÕES ---
PASTA = r"C:\Users\SVOBODA\Desktop\DASHBOARD"
HTML_SRC = os.path.join(PASTA, "dashboard_svoboda_template.html")
HTML_OUT = os.path.join(PASTA, "dashboard_svoboda_atualizado.html")

# --- FUNÇÕES DE APOIO ---
def gv(row, col):
    try:
        if col >= len(row): return ""
        v = row[col]
        return str(v).strip() if v is not None else ""
    except: return ""

def norm_str(s):
    if not s: return ""
    s = str(s).strip()
    return "".join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').title()

def generate_donut_svg(label_counts):
    total = sum(label_counts.values())
    if total == 0: return ""
    colors = {"Concluída": "#22d3a0", "Cancelada": "#ff4d6d", "Não Concluída": "#fbbf24", "Suspensa": "#7c3aed", "Pendente": "#6b7280"}
    svg = f'<svg class="donut-svg" width="120" height="120" viewBox="0 0 42 42">'
    svg += '<circle class="donut-hole" cx="21" cy="21" r="15.915" fill="transparent"></circle>'
    svg += '<circle class="donut-ring" cx="21" cy="21" r="15.915" fill="transparent" stroke="rgba(255,255,255,0.05)" stroke-width="3"></circle>'
    offset = 25
    legend = '<div class="donut-legend">'
    for label, count in label_counts.items():
        if count == 0: continue
        percent = (count / total) * 100
        color = colors.get(label, "#00e5ff")
        dash = f"{percent} {100-percent}"
        svg += f'<circle class="donut-segment" cx="21" cy="21" r="15.915" fill="transparent" stroke="{color}" stroke-width="3" stroke-dasharray="{dash}" stroke-dashoffset="{offset}"></circle>'
        offset -= percent
        legend += f'<div class="legend-item"><div class="legend-dot" style="background:{color}"></div>{label}<div class="legend-val">{count} ({percent:.1f}%)</div></div>'
    svg += '</svg></div>' + legend + '</div>'
    return f'<div class="donut-wrap">{svg}'

def generate_motivos_html(motivos_counter):
    top_motivos = motivos_counter.most_common(6)
    if not top_motivos: return '<div class="chart-card" id="motivos-container"><div class="chart-title">Top Motivos</div><div style="color:var(--muted); font-size:12px; font-family:var(--font-mono); height:200px; display:flex; align-items:center; justify-content:center;">Nenhum motivo registrado</div></div>'
    max_val = top_motivos[0][1]
    html = '<div class="chart-card" id="motivos-container"><div class="chart-title">Top Motivos de Não Conclusão / Cancelamento</div>'
    for motivo, count in top_motivos:
        width = (count / max_val) * 100
        html += f'<div class="bar-row"><div class="bar-label wide">{motivo}</div><div class="bar-track"><div class="bar-fill red" style="width:{width}%"></div></div><div class="bar-val">{count}</div></div>'
    html += '</div>'
    return html

# --- 1. DEFINE A PLANILHA PRINCIPAL DE HISTÓRICO ---
caminho_os = os.path.join(PASTA, "ATIVIDADES DAS O.S.xlsx")
if not os.path.exists(caminho_os):
    print(f"[ERRO] A planilha {caminho_os} nao foi encontrada!")
    exit()

print(f"[OK] Lendo: {os.path.basename(caminho_os)}")

# --- 2. CARREGA DADOS ---
wb = openpyxl.load_workbook(caminho_os, read_only=True, data_only=True)
aba = next((n for n in ["Plan1", "BASE_DADOS", "Base"] if n in wb.sheetnames), wb.sheetnames[0])
ws = wb[aba]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

# Normaliza headers removendo acentos para garantir match independente de encoding
def _norm_hdr(h):
    if not h: return ""
    s = unicodedata.normalize('NFD', str(h).strip())
    return "".join(c for c in s if unicodedata.category(c) != 'Mn')

idx = {_norm_hdr(h): i for i, h in enumerate(headers) if h}

CR       = idx.get("Recurso", 1)
CD       = idx.get("Data", 2)
CS       = idx.get("Status da Atividade", 3)
CC       = idx.get("Cidade", 6)
CT2      = idx.get("Tipo de Atividade_2", idx.get("Tipo de Atividade 2", idx.get("Tipo de Atividade", 24)))
if CT2 == idx.get("Tipo de Atividade") and CT2 is not None:
    # Se achou só o genérico, tenta forçar a coluna 24 como fallback caso ela exista na planilha
    if len(headers) > 24 and str(headers[24]).startswith("Tipo"):
        CT2 = 24
CMOT     = idx.get("Motivo de Encerramento das atividades", 38)
CDUR     = idx.get("Duracao", 21)              # sem acento apos normalizacao
CDESL    = idx.get("Tempo de Deslocamento", 22)
CSLA_FIM = idx.get("Fim do SLA", 20)
COS_NUM  = idx.get("Atividade", idx.get("Numero da Atividade", idx.get("N da Atividade", 4)))
print(f"[DEBUG] Colunas OS: CR={CR} CD={CD} CS={CS} CC={CC} CT2={CT2} CMOT={CMOT} CDUR={CDUR} CDESL={CDESL} CSLA_FIM={CSLA_FIM} COS_NUM={COS_NUM}")

tags_tec, tags_city, tags_tipo, tags_dia = {}, {}, {}, {}
motivos_canc = Counter()
status_counts = {"Concluída": 0, "Cancelada": 0, "Não Concluída": 0, "Suspensa": 0, "Pendente": 0}
raw_os_data = []
tempos_duracao = []
tempos_deslocamento = []
sla_estourado = 0
sla_ativacao_list = []  # Lista de ativações pendentes para seção SLA TV

# Normaliza valor de status removendo acentos para comparacao robusta
def _norm_status(s):
    n = unicodedata.normalize('NFD', str(s).lower())
    return "".join(c for c in n if unicodedata.category(c) != 'Mn')

total_os = 0
for row in rows[1:]:
    status_raw = gv(row, CS)
    if not status_raw: continue
    status_norm = _norm_status(status_raw)
    cat = "Pendente"
    # IMPORTANTE: verificar "nao conclu" ANTES de "conclu" para nao classificar errado
    if "nao conclu" in status_norm:   cat = "Não Concluída"
    elif "conclu" in status_norm:     cat = "Concluída"
    elif "cancel" in status_norm:     cat = "Cancelada"
    elif "suspen" in status_norm:     cat = "Suspensa"
    elif "iniciado" in status_norm or "em rota" in status_norm: cat = "Pendente"
    
    total_os += 1
    status_counts[cat] += 1
    
    recurso = norm_str(gv(row, CR))
    cidade = norm_str(gv(row, CC))
    tipo = gv(row, CT2)
    motivo = gv(row, CMOT)
    data_raw = row[CD] if len(row) > CD else None
    
    data_key = "--"
    if data_raw:
        try:
            dt_obj = data_raw if not isinstance(data_raw, str) else datetime.strptime(str(data_raw)[:10], "%Y-%m-%d")
            data_key = dt_obj.strftime("%d/%m")
        except: pass

    # Coleta de tempos
    dur_raw = gv(row, CDUR)
    desl_raw = gv(row, CDESL)
    dur_min = 0
    desl_min = 0
    if dur_raw:
        try:
            dur_min = int(str(dur_raw).split(':')[0]) * 60 + int(str(dur_raw).split(':')[1]) if ':' in str(dur_raw) else int(float(str(dur_raw)))
            if dur_min > 0: tempos_duracao.append(dur_min)
        except: pass
    if desl_raw:
        try:
            desl_min = int(str(desl_raw).split(':')[0]) * 60 + int(str(desl_raw).split(':')[1]) if ':' in str(desl_raw) else int(float(str(desl_raw)))
            if desl_min > 0: tempos_deslocamento.append(desl_min)
        except: pass
    # SLA: abertura (col 15) até data atividade + hora fim (col 2 + col 17)
    sla_flag = 0
    os_num = gv(row, COS_NUM)
    if cat == "Concluída":
        try:
            abertura = row[15]  # datetime
            data_ativ = row[CD]  # Data da atividade
            hora_fim = row[17]  # time
            if abertura and data_ativ and hora_fim:
                from datetime import timedelta as td
                dt_fim = datetime.combine(data_ativ.date() if hasattr(data_ativ, 'date') else data_ativ, hora_fim)
                diff_h = (dt_fim - abertura).total_seconds() / 3600
                if diff_h < 0:
                    dt_fim = datetime.combine((data_ativ.date() if hasattr(data_ativ, 'date') else data_ativ) + td(days=1), hora_fim)
                    diff_h = (dt_fim - abertura).total_seconds() / 3600
                if diff_h > 24:
                    sla_estourado += 1
                    sla_flag = 1
        except: pass

    # SLA DE ATIVAÇÃO: coletar ativações pendentes com prazo
    if cat == "Pendente" and tipo and "ativa" in tipo.lower():
        try:
            fim_sla = row[CSLA_FIM]
            if fim_sla and hasattr(fim_sla, 'timestamp'):
                diff_sla = (fim_sla - datetime.now()).total_seconds()
                if diff_sla > 0:
                    # No prazo — contagem regressiva
                    sla_ativacao_list.append({"tec": recurso, "tipo": tipo, "os": os_num, "city": cidade, "status_sla": "no_prazo", "restante_min": int(diff_sla / 60)})
                else:
                    # Fora do prazo — para a contagem
                    sla_ativacao_list.append({"tec": recurso, "tipo": tipo, "os": os_num, "city": cidade, "status_sla": "fora", "restante_min": 0})
        except:
            pass

    raw_os_data.append({"tec": recurso, "city": cidade, "status": cat, "type": tipo, "date": data_key, "motivo": motivo, "dur": dur_min, "desl": desl_min, "sla": sla_flag, "os": os_num})
    if cat in ["Cancelada", "Não Concluída"] and motivo: motivos_canc[motivo] += 1

    for d, k in [(tags_tec, recurso), (tags_city, cidade), (tags_tipo, tipo)]:
        if k not in d: d[k] = {"key": k, "total": 0, "concluido": 0, "nao_concluido": 0, "cancelado": 0}
        d[k]["total"] += 1
        if cat == "Concluída": d[k]["concluido"] += 1
        elif cat == "Não Concluída": d[k]["nao_concluido"] += 1
        elif cat == "Cancelada": d[k]["cancelado"] += 1

    if data_key != "--":
        if data_key not in tags_dia: tags_dia[data_key] = {"data": data_key, "total": 0, "concluido": 0}
        tags_dia[data_key]["total"] += 1
        if cat == "Concluída": tags_dia[data_key]["concluido"] += 1

def calc_taxa(d):
    d["taxa"] = round((d["concluido"] / d["total"] * 100), 1) if d["total"] > 0 else 0
    if "key" in d:
        if d["key"] in tags_tec: d["Recurso"] = d["key"]
        elif d["key"] in tags_city: d["Cidade"] = d["key"]
        elif d["key"] in tags_tipo: d["tipo"] = d["key"]
    return d

list_tec = sorted([calc_taxa(v) for v in tags_tec.values()], key=lambda x: x["taxa"], reverse=True)
list_city = sorted([calc_taxa(v) for v in tags_city.values()], key=lambda x: x["total"], reverse=True)
list_tipo = sorted([calc_taxa(v) for v in tags_tipo.values()], key=lambda x: x["total"], reverse=True)
list_diag = sorted([calc_taxa(v) for v in tags_dia.values()], key=lambda x: x["data"])

# Recordes
v_tec_taxa = list_tec[0] if list_tec else {"Recurso": "-", "taxa": 0}
v_tec_prod = sorted(list_tec, key=lambda x: x["total"], reverse=True)[0] if list_tec else {"Recurso": "-", "total": 0}
v_tec_bad = list_tec[-1] if list_tec else {"Recurso": "-", "taxa": 0}
v_cit_taxa = sorted(list_city, key=lambda x: x["taxa"], reverse=True)[0] if list_city else {"Cidade": "-", "taxa": 0}
v_cit_vol = list_city[0] if list_city else {"Cidade": "-", "total": 0}
v_cit_bad = sorted(list_city, key=lambda x: x["taxa"])[0] if list_city else {"Cidade": "-", "taxa": 0}
v_pico = sorted(list_diag, key=lambda x: x["total"], reverse=True)[0] if list_diag else {"data": "-", "total": 0}
v_melhor = sorted(list_diag, key=lambda x: x["taxa"], reverse=True)[0] if list_diag else {"data": "-", "taxa": 0}
v_pior = sorted(list_diag, key=lambda x: x["taxa"])[0] if list_diag else {"data": "-", "taxa": 0}
v_media = total_os / len(list_diag) if list_diag else 0

# --- 2b. DOWNLOAD E LEITURA DA PLANILHA DE FROTA ---
FROTA_FILE = os.path.join(PASTA, "frota_temp.xlsx")
FROTA_SHEET_ID = "1CjWOfMTQQ1aP3jeF-Wf3gxIa5DT0qMRwjijD6nXXtQg"
FROTA_URL = f"https://docs.google.com/spreadsheets/d/{FROTA_SHEET_ID}/export?format=xlsx"

# Download automático da planilha de frota
try:
    import urllib.request
    frota_temp_dl = os.path.join(PASTA, "frota_download.xlsx")
    print("[...] Baixando planilha de frota do Google Sheets...")
    urllib.request.urlretrieve(FROTA_URL, frota_temp_dl)
    print("[OK] Planilha de frota baixada!")
    # Tentar substituir o arquivo principal
    try:
        import shutil
        shutil.move(frota_temp_dl, FROTA_FILE)
        print("[OK] frota_temp.xlsx atualizado!")
    except:
        FROTA_FILE = frota_temp_dl
        print("[INFO] Usando arquivo baixado diretamente")
except Exception as e:
    print(f"[AVISO] Não foi possível baixar a frota: {e}")
    print("[INFO] Usando arquivo local existente (se houver)")

frota_veiculos = []
frota_oficina = []
frota_revisao = []

if os.path.exists(FROTA_FILE):
    print(f"[OK] Lendo frota: frota_temp.xlsx")
    wb_frota = openpyxl.load_workbook(FROTA_FILE, read_only=True, data_only=True)
    
    # PLACA X TÉCNICOS
    if "PLACA X TÉCNICOS" in wb_frota.sheetnames:
        ws_f = wb_frota["PLACA X TÉCNICOS"]
        for r in list(ws_f.iter_rows(values_only=True))[1:]:
            placa = str(r[0]).strip() if r[0] else ""
            func = str(r[1]).strip() if r[1] else ""
            locadora = str(r[3]).strip() if r[3] else ""
            status = str(r[4]).strip() if r[4] else ""
            if placa and placa != "None":
                frota_veiculos.append({"placa": placa, "tecnico": func, "locadora": locadora, "status": status})
    
    # OFICINA
    oficina_nome = next((n for n in wb_frota.sheetnames if "OFICINA" in n.upper()), None)
    if oficina_nome:
        ws_o = wb_frota[oficina_nome]
        for r in list(ws_o.iter_rows(values_only=True))[1:]:
            status_o = str(r[6]).strip() if r[6] else ""
            if status_o and status_o.upper() not in ["N/A", "None", "PRONTO", "RETIRADO", ""] and "OFICINA" in status_o.upper():
                placa = str(r[0]).strip() if r[0] else ""
                motorista = str(r[1]).strip() if r[1] else ""
                entrada_raw = r[3]
                entrada = str(r[3])[:10] if r[3] else "-"
                previsao = str(r[4])[:10] if r[4] else "-"
                saida_raw = r[5]
                cidade = str(r[7]).strip() if r[7] else ""
                obs = str(r[8]).strip() if r[8] else ""
                # Calcular dias na oficina
                dias = 0
                try:
                    from datetime import timedelta as td2
                    dt_entrada = entrada_raw if hasattr(entrada_raw, 'date') else datetime.strptime(str(entrada_raw)[:10], '%Y-%m-%d')
                    if saida_raw and str(saida_raw).strip() not in ['', 'None', '-']:
                        dt_saida = saida_raw if hasattr(saida_raw, 'date') else datetime.strptime(str(saida_raw)[:10], '%Y-%m-%d')
                    else:
                        dt_saida = datetime.now()
                    dias = (dt_saida - dt_entrada).days
                    if dias < 0: dias = 0
                except: pass
                if placa and placa != "None":
                    frota_oficina.append({"placa": placa, "motorista": motorista, "entrada": entrada, "previsao": previsao, "status": status_o, "cidade": cidade, "obs": obs, "dias": dias})
    
    # ANÁLISE DE REVISÃO
    revisao_nome = next((n for n in wb_frota.sheetnames if "REVIS" in n.upper()), None)
    if revisao_nome:
        ws_r = wb_frota[revisao_nome]
        for r in list(ws_r.iter_rows(values_only=True))[1:]:
            rev_status = str(r[5]).strip() if r[5] else ""
            if rev_status in ["REVISÃO", "AGENDAR", "REVISAR"]:
                tec = str(r[0]).strip() if r[0] else ""
                placa = str(r[1]).strip() if r[1] else ""
                if placa.upper() == "PLACA": continue  # subtítulo de cidade
                locadora = str(r[2]).strip() if r[2] else ""
                km = str(r[3]).split(".")[0] if r[3] else "-"
                data_rev = str(r[4])[:10] if r[4] else "-"
                if placa and placa != "None":
                    frota_revisao.append({"tecnico": tec, "placa": placa, "locadora": locadora, "km": km, "data": data_rev, "status": rev_status})
    
    # COMBUSTÍVEL $
    comb_nome = next((n for n in wb_frota.sheetnames if "COMBUST" in n.upper()), None)
    frota_combustivel = []
    total_gasto_comb = 0
    total_litros_comb = 0
    total_km_comb = 0
    if comb_nome:
        ws_c = wb_frota[comb_nome]
        for r in list(ws_c.iter_rows(values_only=True)):
            nome = str(r[0]).strip() if r[0] else ""
            if not nome or nome in ["-", "None", "l"]: continue
            if "equipe" in nome.lower() or "BH" in nome or "=" in nome or "Home" in nome or "ALMOX" in nome.upper(): continue
            try:
                soma = float(str(r[4]).replace("R$","").replace(",",".").strip()) if r[4] and str(r[4]).strip() not in ["-", "None", "SOMA", "SOMA "] else 0
                litros = float(r[5]) if r[5] and str(r[5]).strip() not in ["-", "None", "LITROS", "LITROS "] else 0
                km = float(r[6]) if r[6] and str(r[6]).strip() not in ["-", "None", "KM", "KM "] else 0
            except: soma = litros = km = 0
            if soma > 0:
                frota_combustivel.append({"tecnico": nome, "gasto": soma, "litros": round(litros, 1), "km": round(km, 0)})
                total_gasto_comb += soma
                total_litros_comb += litros
                total_km_comb += km
    frota_combustivel.sort(key=lambda x: x["gasto"], reverse=True)
    
    # Projeção 30 dias
    dia_atual = datetime.now().day
    projecao_30d = (total_gasto_comb / dia_atual) * 30 if dia_atual > 0 else 0
    
    # ACOMPANHAMENTO DE KM (km rodado real - da aba MAIO ou ABRIL, coluna 32)
    km_rodado_por_tec = {}
    mes_nome = next((n for n in wb_frota.sheetnames if "MAIO" in n.upper()), None)
    if not mes_nome:
        mes_nome = next((n for n in wb_frota.sheetnames if "ABRIL" in n.upper()), None)
    if mes_nome:
        ws_mes = wb_frota[mes_nome]
        for r in list(ws_mes.iter_rows(values_only=True)):
            nome_km = str(r[0]).strip().upper() if r[0] else ""
            if not nome_km or "EQUIPE" in nome_km or "=" in nome_km or "HOME" in nome_km or "ALMOX" in nome_km: continue
            if len(r) > 32 and r[32]:
                try:
                    km_val = float(r[32])
                    if km_val > 0: km_rodado_por_tec[nome_km] = km_val
                except: pass
    
    # Cruzar KM rodado com combustivel
    for c in frota_combustivel:
        nome_comb = c["tecnico"].strip().upper()
        c["km_rodado"] = 0
        # Match exato
        if nome_comb in km_rodado_por_tec:
            c["km_rodado"] = km_rodado_por_tec[nome_comb]
        else:
            # Match parcial (primeiras 3 palavras)
            palavras_comb = nome_comb.split()[:3]
            for tec_km, km_val in km_rodado_por_tec.items():
                palavras_km = tec_km.split()[:3]
                if len(palavras_comb) >= 2 and len(palavras_km) >= 2:
                    if palavras_comb[0] == palavras_km[0] and palavras_comb[1] == palavras_km[1]:
                        c["km_rodado"] = km_val
                        break
    
    # DESLOCAMENTO
    desl_nome = next((n for n in wb_frota.sheetnames if "DESLOCAMENTO" in n.upper()), None)
    frota_deslocamento = []
    desl_por_tec = {}
    desl_por_dest = {}
    total_km_desl = 0
    total_val_desl = 0
    if desl_nome:
        ws_d = wb_frota[desl_nome]
        for r in list(ws_d.iter_rows(values_only=True))[1:]:
            tec_d = str(r[1]).strip() if r[1] else ""
            dest = str(r[2]).strip() if r[2] else ""
            try: km_d = float(r[3])
            except: km_d = 0
            try: val_d = float(r[4])
            except: val_d = 0
            if tec_d and tec_d != "None":
                if tec_d not in desl_por_tec: desl_por_tec[tec_d] = {"viagens": 0, "km": 0, "valor": 0}
                desl_por_tec[tec_d]["viagens"] += 1
                desl_por_tec[tec_d]["km"] += km_d
                desl_por_tec[tec_d]["valor"] += val_d
            if dest and dest != "None":
                if dest not in desl_por_dest: desl_por_dest[dest] = {"viagens": 0, "km": 0}
                desl_por_dest[dest]["viagens"] += 1
                desl_por_dest[dest]["km"] += km_d
            total_km_desl += km_d
            total_val_desl += val_d
    desl_tec_list = sorted(desl_por_tec.items(), key=lambda x: x[1]["km"], reverse=True)
    desl_dest_list = sorted(desl_por_dest.items(), key=lambda x: x[1]["viagens"], reverse=True)
    
    wb_frota.close()
else:
    print("[AVISO] frota_temp.xlsx não encontrado, aba Frota ficará sem dados")
    frota_combustivel = []
    total_gasto_comb = total_litros_comb = total_km_comb = 0
    desl_por_tec = {}
    desl_por_dest = {}
    desl_tec_list = []
    desl_dest_list = []
    total_km_desl = total_val_desl = 0

# KPIs Frota
frota_total = len(frota_veiculos)
frota_ativos = sum(1 for v in frota_veiculos if "ATIVIDADE" in v["status"].upper())
frota_oficina_count = sum(1 for o in frota_oficina if "OFICINA" in o["status"].upper())
frota_revisao_count = len(frota_revisao)
frota_parados = sum(1 for v in frota_veiculos if v["status"].upper() in ["PARADO", "DESMOBILIZADO", "DEVOLVIDO", "ATESTADO"])

# Gerar HTML da tabela de veículos
def gen_frota_table(veiculos):
    rows_html = ""
    for v in veiculos:
        st = v["status"].upper()
        if "ATIVIDADE" in st: badge = '<span class="badge green">Em Atividade</span>'
        elif "OFICINA" in st: badge = '<span class="badge red">Oficina</span>'
        elif st in ["PARADO", "DESMOBILIZADO"]: badge = '<span class="badge yellow">Parado</span>'
        elif "DEVOLVIDO" in st: badge = '<span class="badge yellow">Devolvido</span>'
        elif "ATESTADO" in st: badge = '<span class="badge yellow">Atestado</span>'
        else: badge = f'<span class="badge">{v["status"]}</span>'
        rows_html += f'<tr><td style="font-weight:600;color:#e8eaf6">{v["placa"]}</td><td>{v["tecnico"]}</td><td>{v["locadora"]}</td><td>{badge}</td></tr>\n'
    return rows_html

def gen_oficina_table(oficina):
    rows_html = ""
    for o in sorted(oficina, key=lambda x: x.get('dias', 0), reverse=True):
        st = o["status"].upper()
        if "OFICINA" in st: badge = '<span class="badge red">Na Oficina</span>'
        elif "AGEND" in st: badge = '<span class="badge yellow">Agendado</span>'
        else: badge = f'<span class="badge">{o["status"]}</span>'
        dias = o.get('dias', 0)
        if dias > 30: dias_badge = f'<span class="badge red">{dias}d</span>'
        elif dias > 14: dias_badge = f'<span class="badge yellow">{dias}d</span>'
        else: dias_badge = f'<span class="badge green">{dias}d</span>'
        rows_html += f'<tr><td style="font-weight:600;color:#e8eaf6">{o["placa"]}</td><td>{o["motorista"]}</td><td>{o["entrada"]}</td><td>{o["previsao"]}</td><td>{o["cidade"]}</td><td>{badge}</td><td>{dias_badge}</td></tr>\n'
    return rows_html

def gen_revisao_table(revisao):
    rows_html = ""
    for r in revisao:
        st = r["status"].upper()
        if "AGENDAR" in st: badge = '<span class="badge yellow">Agendar</span>'
        else: badge = '<span class="badge red">Revisão</span>'
        rows_html += f'<tr><td style="font-weight:600;color:#e8eaf6">{r["placa"]}</td><td>{r["tecnico"]}</td><td>{r["locadora"]}</td><td>{r["km"]} km</td><td>{r["data"]}</td><td>{badge}</td></tr>\n'
    return rows_html

def gen_combustivel_table(combustivel):
    rows_html = ""
    for i, c in enumerate(combustivel):
        km_rodado = c.get("km_rodado", 0)
        km_esperado = c["km"]
        pct = round((km_rodado / km_esperado) * 100) if km_esperado > 0 else 0
        pct = min(pct, 100)
        if pct >= 80: bar_color = "#22d3a0"
        elif pct >= 50: bar_color = "#fbbf24"
        else: bar_color = "#ff4d6d"
        bar_html = f'<div style="display:flex;align-items:center;gap:10px;"><div style="flex:1;background:rgba(255,255,255,0.06);border-radius:4px;height:10px;min-width:120px;"><div style="width:{pct}%;height:100%;background:{bar_color};border-radius:4px;transition:width 0.3s;"></div></div><span style="font-size:12px;color:{bar_color};font-weight:700;min-width:40px;text-align:right;">{pct}%</span></div>'
        rows_html += f'<tr><td style="padding:10px 8px;text-align:center;">{i+1}</td><td style="padding:10px 12px;">{c["tecnico"]}</td><td style="padding:10px 12px;color:#f472b6;font-weight:600;text-align:right;">R$ {c["gasto"]:,.2f}</td><td style="padding:10px 12px;text-align:right;">{c["litros"]}</td><td style="padding:10px 12px;text-align:right;color:#60a5fa;font-weight:600;">{km_rodado:,.0f}</td><td style="padding:10px 12px;text-align:right;color:#94a3b8;">{km_esperado:,.0f}</td><td style="padding:10px 12px;min-width:180px;">{bar_html}</td></tr>\n'
    return rows_html

def gen_desl_tec_table(tec_list):
    rows_html = ""
    for i, t in enumerate(tec_list):
        rows_html += f'<tr><td>{i+1}</td><td>{t[0]}</td><td>{t[1]["viagens"]}</td><td>{t[1]["km"]:,.0f}</td><td>R$ {t[1]["valor"]:,.2f}</td></tr>\n'
    return rows_html

def gen_desl_dest_table(dest_list):
    rows_html = ""
    for i, d in enumerate(dest_list):
        rows_html += f'<tr><td>{i+1}</td><td style="font-weight:600;color:#e8eaf6">{d[0]}</td><td>{d[1]["viagens"]}</td><td>{d[1]["km"]:,.0f}</td></tr>\n'
    return rows_html

frota_page_html = f"""
<!-- ==================== FROTA ==================== -->
<style>
  .frota-tabs {{ display:flex; gap:8px; margin-bottom:24px; }}
  .frota-tab {{ background:#111520; border:1px solid #1c2237; border-radius:8px; padding:10px 20px; font-size:13px; color:#94a3b8; cursor:pointer; font-family:var(--font-head); font-weight:600; transition:all 0.2s; }}
  .frota-tab:hover {{ border-color:#7c3aed; color:#e8eaf6; }}
  .frota-tab.active {{ background:#7c3aed; border-color:#7c3aed; color:white; box-shadow:0 4px 15px rgba(124,58,237,0.3); }}
  .frota-sub {{ display:none; }}
  .frota-sub.active {{ display:block; }}
</style>
<div class="page" id="frota">

  <div class="frota-tabs">
    <div class="frota-tab active" onclick="showFrotaSub('frota-veiculos', this)">🚗 Veículos</div>
    <div class="frota-tab" onclick="showFrotaSub('frota-manut', this)">🔧 Manutenção</div>
    <div class="frota-tab" onclick="showFrotaSub('frota-comb', this)">⛽ Combustível</div>
    <div class="frota-tab" onclick="showFrotaSub('frota-desl', this)">🛣️ Deslocamento</div>
  </div>

  <!-- SUB-ABA: Veículos -->
  <div class="frota-sub active" id="frota-veiculos">
    <div class="kpi-grid" style="grid-template-columns:repeat(5,1fr);margin-bottom:24px;">
      <div class="kpi-card blue">
        <div class="kpi-label">Total Veículos</div>
        <div class="kpi-value blue">{frota_total}</div>
        <div class="kpi-sub">frota completa</div>
      </div>
      <div class="kpi-card green">
        <div class="kpi-label">Em Atividade</div>
        <div class="kpi-value green">{frota_ativos}</div>
        <div class="kpi-sub">{round(frota_ativos/frota_total*100) if frota_total else 0}% da frota</div>
      </div>
      <div class="kpi-card red">
        <div class="kpi-label">Em Oficina</div>
        <div class="kpi-value red">{frota_oficina_count}</div>
        <div class="kpi-sub">manutenção ativa</div>
      </div>
      <div class="kpi-card yellow">
        <div class="kpi-label">Revisão Pendente</div>
        <div class="kpi-value yellow">{frota_revisao_count}</div>
        <div class="kpi-sub">necessitam agendar</div>
      </div>
      <div class="kpi-card">
        <div class="kpi-label">Parados/Outros</div>
        <div class="kpi-value white">{frota_parados}</div>
        <div class="kpi-sub">indisponíveis</div>
      </div>
    </div>
    <div class="section-title">🚗 Veículos da Frota ({frota_total})</div>
    <div class="table-wrap" style="max-height:500px;overflow-y:auto;">
      <table class="data-table" id="tbl-veiculos">
        <thead><tr>
          <th class="sortable" onclick="sortTable('tbl-veiculos',0,'str')">Placa</th>
          <th class="sortable" onclick="sortTable('tbl-veiculos',1,'str')">Técnico</th>
          <th class="sortable" onclick="sortTable('tbl-veiculos',2,'str')">Locadora</th>
          <th class="sortable" onclick="sortTable('tbl-veiculos',3,'str')">Status</th>
        </tr></thead>
        <tbody>{gen_frota_table(frota_veiculos)}</tbody>
      </table>
    </div>
  </div>

  <!-- SUB-ABA: Manutenção -->
  <div class="frota-sub" id="frota-manut">
    <div class="kpi-grid" style="grid-template-columns:repeat(2,1fr);margin-bottom:20px;">
      <div class="kpi-card red">
        <div class="kpi-label">Em Oficina</div>
        <div class="kpi-value red">{len(frota_oficina)}</div>
        <div class="kpi-sub">veículos em manutenção</div>
      </div>
      <div class="kpi-card yellow">
        <div class="kpi-label">Revisão Pendente</div>
        <div class="kpi-value yellow">{frota_revisao_count}</div>
        <div class="kpi-sub">necessitam agendar</div>
      </div>
    </div>
    <div class="grid-2">
      <div>
        <div class="section-title">🔧 Veículos em Oficina ({len(frota_oficina)})</div>
        <div class="table-wrap" style="max-height:400px;overflow-y:auto;">
          <table class="data-table" id="tbl-oficina">
            <thead><tr>
              <th class="sortable" onclick="sortTable('tbl-oficina',0,'str')">Placa</th>
              <th class="sortable" onclick="sortTable('tbl-oficina',1,'str')">Motorista</th>
              <th class="sortable" onclick="sortTable('tbl-oficina',2,'str')">Entrada</th>
              <th class="sortable" onclick="sortTable('tbl-oficina',3,'str')">Previsão</th>
              <th class="sortable" onclick="sortTable('tbl-oficina',4,'str')">Cidade</th>
              <th class="sortable" onclick="sortTable('tbl-oficina',5,'str')">Status</th>
              <th class="sortable desc" onclick="sortTable('tbl-oficina',6,'num')">Dias</th>
            </tr></thead>
            <tbody>{gen_oficina_table(frota_oficina)}</tbody>
          </table>
        </div>
      </div>
      <div>
        <div class="section-title">📋 Revisões Pendentes ({frota_revisao_count})</div>
        <div class="table-wrap" style="max-height:400px;overflow-y:auto;">
          <table class="data-table" id="tbl-revisao">
            <thead><tr>
              <th class="sortable" onclick="sortTable('tbl-revisao',0,'str')">Placa</th>
              <th class="sortable" onclick="sortTable('tbl-revisao',1,'str')">Técnico</th>
              <th class="sortable" onclick="sortTable('tbl-revisao',2,'str')">Locadora</th>
              <th class="sortable" onclick="sortTable('tbl-revisao',3,'num')">KM</th>
              <th class="sortable" onclick="sortTable('tbl-revisao',4,'str')">Data</th>
              <th class="sortable" onclick="sortTable('tbl-revisao',5,'str')">Status</th>
            </tr></thead>
            <tbody>{gen_revisao_table(frota_revisao)}</tbody>
          </table>
        </div>
      </div>
    </div>
  </div>

  <!-- SUB-ABA: Combustível -->
  <div class="frota-sub" id="frota-comb">
    <div class="kpi-grid" style="grid-template-columns:repeat(4,1fr);margin-bottom:20px;">
      <div class="kpi-card red">
        <div class="kpi-label">Total Gasto</div>
        <div class="kpi-value red">R$ {total_gasto_comb:,.2f}</div>
        <div class="kpi-sub">{len(frota_combustivel)} técnicos</div>
      </div>
      <div class="kpi-card" style="border-color:#7c3aed33;">
        <div class="kpi-label">📊 Projeção 30 Dias</div>
        <div class="kpi-value" style="color:#a78bfa;">R$ {projecao_30d:,.2f}</div>
        <div class="kpi-sub">projeção mensal (dia {dia_atual}/30)</div>
      </div>
      <div class="kpi-card blue">
        <div class="kpi-label">Litros Abastecidos</div>
        <div class="kpi-value blue">{total_litros_comb:,.0f}L</div>
        <div class="kpi-sub">média {round(total_litros_comb/len(frota_combustivel)) if frota_combustivel else 0}L/técnico</div>
      </div>
      <div class="kpi-card green">
        <div class="kpi-label">KM Estimado</div>
        <div class="kpi-value green">{total_km_comb:,.0f}</div>
        <div class="kpi-sub">média {round(total_km_comb/len(frota_combustivel)) if frota_combustivel else 0} km/técnico</div>
      </div>
    </div>
    <div class="section-title">⛽ Ranking de Combustível</div>
    <div class="table-wrap" style="max-height:500px;overflow-y:auto;">
      <table class="data-table" id="tbl-comb">
        <thead><tr>
          <th class="sortable" onclick="sortTable('tbl-comb',0,'num')" style="padding:10px 8px;">#</th>
          <th class="sortable" onclick="sortTable('tbl-comb',1,'str')" style="padding:10px 12px;">Técnico</th>
          <th class="sortable desc" onclick="sortTable('tbl-comb',2,'num')" style="padding:10px 12px;text-align:right;">Gasto (R$)</th>
          <th class="sortable" onclick="sortTable('tbl-comb',3,'num')" style="padding:10px 12px;text-align:right;">Litros</th>
          <th class="sortable" onclick="sortTable('tbl-comb',4,'num')" style="padding:10px 12px;text-align:right;">KM Rodado</th>
          <th class="sortable" onclick="sortTable('tbl-comb',5,'num')" style="padding:10px 12px;text-align:right;">KM Esperado</th>
          <th style="padding:10px 12px;">Progresso</th>
        </tr></thead>
        <tbody>{gen_combustivel_table(frota_combustivel)}</tbody>
      </table>
    </div>
  </div>

  <!-- SUB-ABA: Deslocamento -->
  <div class="frota-sub" id="frota-desl">
    <div class="kpi-grid" style="grid-template-columns:repeat(3,1fr);margin-bottom:20px;">
      <div class="kpi-card blue">
        <div class="kpi-label">KM Total</div>
        <div class="kpi-value blue">{total_km_desl:,.0f}</div>
        <div class="kpi-sub">{len(desl_tec_list)} técnicos</div>
      </div>
      <div class="kpi-card red">
        <div class="kpi-label">Custo Total</div>
        <div class="kpi-value red">R$ {total_val_desl:,.2f}</div>
        <div class="kpi-sub">R$ {round(total_val_desl/sum(v["viagens"] for _,v in desl_tec_list)) if desl_tec_list else 0:.0f}/viagem</div>
      </div>
      <div class="kpi-card green">
        <div class="kpi-label">Total Viagens</div>
        <div class="kpi-value green">{sum(v["viagens"] for _,v in desl_tec_list)}</div>
        <div class="kpi-sub">{len(desl_dest_list)} destinos</div>
      </div>
    </div>
    <div class="grid-2">
      <div>
        <div class="section-title">🛣️ Por Técnico</div>
        <div class="table-wrap" style="max-height:400px;overflow-y:auto;">
          <table class="data-table" id="tbl-desl-tec">
            <thead><tr>
              <th class="sortable" onclick="sortTable('tbl-desl-tec',0,'num')">#</th>
              <th class="sortable" onclick="sortTable('tbl-desl-tec',1,'str')">Técnico</th>
              <th class="sortable" onclick="sortTable('tbl-desl-tec',2,'num')">Viagens</th>
              <th class="sortable desc" onclick="sortTable('tbl-desl-tec',3,'num')">KM Total</th>
              <th class="sortable" onclick="sortTable('tbl-desl-tec',4,'num')">Valor (R$)</th>
            </tr></thead>
            <tbody>{gen_desl_tec_table(desl_tec_list)}</tbody>
          </table>
        </div>
      </div>
      <div>
        <div class="section-title">📍 Top Destinos</div>
        <div class="table-wrap" style="max-height:400px;overflow-y:auto;">
          <table class="data-table" id="tbl-desl-dest">
            <thead><tr>
              <th class="sortable" onclick="sortTable('tbl-desl-dest',0,'num')">#</th>
              <th class="sortable" onclick="sortTable('tbl-desl-dest',1,'str')">Destino</th>
              <th class="sortable desc" onclick="sortTable('tbl-desl-dest',2,'num')">Viagens</th>
              <th class="sortable" onclick="sortTable('tbl-desl-dest',3,'num')">KM Total</th>
            </tr></thead>
            <tbody>{gen_desl_dest_table(desl_dest_list)}</tbody>
          </table>
        </div>
      </div>
    </div>
  </div>

  <script>
  function showFrotaSub(id, btn) {{
    document.querySelectorAll('.frota-sub').forEach(function(s) {{ s.classList.remove('active'); }});
    document.querySelectorAll('.frota-tab').forEach(function(b) {{ b.classList.remove('active'); }});
    document.getElementById(id).classList.add('active');
    if(btn) btn.classList.add('active');
  }}

  /* ── Ordenação genérica para tabelas da Frota ── */
  var _frotaSortState = {{}};
  function sortTable(tableId, colIdx, tipo) {{
    var tbl = document.getElementById(tableId);
    if (!tbl) return;
    var state = _frotaSortState[tableId] || {{col:-1, dir:'asc'}};
    var dir = (state.col === colIdx && state.dir === 'desc') ? 'asc' : 'desc';
    _frotaSortState[tableId] = {{col: colIdx, dir: dir}};

    /* atualiza ícones */
    tbl.querySelectorAll('th.sortable').forEach(function(th, i) {{
      th.classList.remove('asc','desc');
      if (i === colIdx) th.classList.add(dir);
    }});

    /* ordena linhas */
    var tbody = tbl.tBodies[0];
    var rows = Array.from(tbody.rows);
    rows.sort(function(a, b) {{
      var va = a.cells[colIdx] ? a.cells[colIdx].innerText.trim() : '';
      var vb = b.cells[colIdx] ? b.cells[colIdx].innerText.trim() : '';
      if (tipo === 'num') {{
        va = parseFloat(va.replace(/[^0-9.,\-]/g,'').replace(',','.')) || 0;
        vb = parseFloat(vb.replace(/[^0-9.,\-]/g,'').replace(',','.')) || 0;
        return dir === 'asc' ? va - vb : vb - va;
      }}
      return dir === 'asc' ? va.localeCompare(vb,'pt') : vb.localeCompare(va,'pt');
    }});
    rows.forEach(function(r) {{ tbody.appendChild(r); }});
  }}
  </script>
</div>
"""

# --- 2c. DOWNLOAD E LEITURA DO PAINEL GIGA+ (INDICADORES) ---
GIGA_SHEET_ID = "1cWX9jvJ1WaAFDkWi9aulZ675tMxEkZrmqdnZSYwcAWc"
GIGA_URL = f"https://docs.google.com/spreadsheets/d/{GIGA_SHEET_ID}/export?format=xlsx"
GIGA_FILE = os.path.join(PASTA, "giga_temp.xlsx")

try:
    import urllib.request, shutil
    giga_dl = os.path.join(PASTA, "giga_download.xlsx")
    print("[...] Baixando painel GIGA+ do Google Sheets...")
    urllib.request.urlretrieve(GIGA_URL, giga_dl)
    shutil.move(giga_dl, GIGA_FILE)
    print("[OK] giga_temp.xlsx atualizado!")
except Exception as e:
    print(f"[AVISO] N\u00e3o foi poss\u00edvel baixar o GIGA+: {e}")

indicadores_data = []
if os.path.exists(GIGA_FILE):
    wb_giga = openpyxl.load_workbook(GIGA_FILE, read_only=True, data_only=True)
    aba_giga = next((n for n in wb_giga.sheetnames if n.upper().strip() == "PAINEL GIGA+"), None)
    if not aba_giga:
        aba_giga = next((n for n in wb_giga.sheetnames if "PAINEL GIGA" in n.upper()), None)
    print(f"[DEBUG] Aba selecionada: '{aba_giga}' (abas: {wb_giga.sheetnames})")
    if aba_giga:
        ws_gi = wb_giga[aba_giga]
        giga_rows = list(ws_gi.iter_rows(values_only=True))

        # Linha 1 (índice 1) = headers específicos das colunas
        headers_row = [str(h).strip().upper() if h else "" for h in giga_rows[1]]
        print(f"[DEBUG] Headers GIGA+: {headers_row}")

        # Mapeamento dinâmico por nome — prefere match exato, depois substring
        def _col(nome, alternativas=[]):
            # 1ª passada: match exato (== em vez de "in")
            for h in [nome] + alternativas:
                for i, hdr in enumerate(headers_row):
                    if hdr == h.upper():
                        return i
            # 2ª passada: substring, mas exige que o header comece com o termo
            for h in [nome] + alternativas:
                for i, hdr in enumerate(headers_row):
                    if hdr.startswith(h.upper()):
                        return i
            # 3ª passada: substring "in" (fallback)
            for h in [nome] + alternativas:
                for i, hdr in enumerate(headers_row):
                    if h.upper() in hdr:
                        return i
            return None

        # Coluna 0 é sempre o nome do técnico (header pode ser "TÉCNICO" ou "COUNTA...")
        C_TEC    = 0
        # Colunas 0-6 podem ter headers vazios por células mescladas no Google Sheets
        # Fallback para posições fixas conhecidas da planilha PAINEL GIGA+
        C_ATV    = _col("ATIVAÇÃO", ["ATIVACAO", "ATIVACÃO"])
        if C_ATV is None: C_ATV = 1
        C_COMODO = _col("MUD. DE CÔMODO", ["CÔMODO", "COMODO", "MUD. DE COMODO"])
        if C_COMODO is None: C_COMODO = 2
        C_ENDER  = _col("MUD. DE ENDEREÇO", ["ENDEREÇO", "ENDERECO", "MUD. DE ENDERECO"])
        if C_ENDER is None: C_ENDER = 3
        C_REP    = _col("REPARO")
        if C_REP is None: C_REP = 4
        C_UPG    = _col("UPGRADE")
        if C_UPG is None: C_UPG = 5
        C_TOT    = _col("TOTAL GERAL", ["TOTAL"])
        if C_TOT is None: C_TOT = 6
        C_MPROD  = _col("MÉDIA PRODUTIVA", ["MEDIA PRODUTIVA"])
        C_PONT   = _col("PONTUAÇÃO", ["PONTUACAO"])
        C_QATV   = _col("QNT IFI ATV")
        C_QMDE   = _col("QNT IFI MDE")
        C_QIRR   = _col("QNT IRR")
        C_NATV   = _col("NOTA IFI ATV")
        C_NMDE   = _col("NOTA IFI MDE")
        C_NIRR   = _col("NOTA IRR")
        C_POS    = _col("PONTUAÇÃO PÓS", ["PONT POS", "PONTUACAO POS"])
        C_COMB   = _col("COMBUSTÍVEL", ["COMBUSTIVEL"])

        print(f"[DEBUG] Índices: TEC={C_TEC} ATV={C_ATV} TOT={C_TOT} MPROD={C_MPROD} PONT={C_PONT} NIRR={C_NIRR} COMB={C_COMB}")

        def _num(v, default=0):
            try:
                s = str(v).replace("R$","").replace(" ","").strip()
                if "," in s:  # Formato BR com milhar: '1.114,55' → remove ponto milhar, troca virgula
                    s = s.replace(".","").replace(",",".")
                return float(s)
            except: return default

        def _pct(v):
            try:
                s = str(v).strip()
                if "%" in s: return round(float(s.replace("%","").replace(",",".").strip()), 1)
                f = float(s.replace(",","."))
                return round(f * 100, 1) if f <= 1.0 else round(f, 1)
            except: return 0.0

        def _get(r, col, default=None):
            if col is None or col >= len(r): return default
            return r[col]

        # --- Lê aba RT para cruzar combustível por técnico ---
        combustivel_map = {}
        aba_rt = next((n for n in wb_giga.sheetnames if n.upper().strip() == "RT"), None)
        if aba_rt:
            ws_rt = wb_giga[aba_rt]
            rt_rows = list(ws_rt.iter_rows(values_only=True))
            print(f"[DEBUG] Aba RT encontrada: '{aba_rt}', {len(rt_rows)} linhas")
            if rt_rows:
                print(f"[DEBUG] RT headers: {rt_rows[0]}")
            # Descobre col do técnico (C=índice 2) e col do valor (D=3 ou E=4)
            rt_headers = [str(h).strip().upper() if h else "" for h in (rt_rows[0] if rt_rows else [])]
            RT_TEC = next((i for i, h in enumerate(rt_headers) if "TÉCNICO" in h or "TECNICO" in h or "NOME" in h), 2)
            RT_VAL = next((i for i, h in enumerate(rt_headers) if "VALOR" in h or "TOTAL" in h or "R$" in h or "COMB" in h), 4)
            print(f"[DEBUG] RT: col_tec={RT_TEC} col_val={RT_VAL}")
            for rr in rt_rows[1:]:
                if len(rr) <= max(RT_TEC, RT_VAL): continue
                nome_rt = str(rr[RT_TEC]).strip() if rr[RT_TEC] else ""
                val_rt  = rr[RT_VAL]
                if nome_rt and nome_rt.lower() not in ["none", "", "técnico", "tecnico"]:
                    try:
                        combustivel_map[nome_rt.upper()] = round(float(str(val_rt).replace(",",".").replace("R$","").replace(" ","").strip()), 2)
                    except:
                        pass
            print(f"[DEBUG] combustível_map: {combustivel_map}")
        else:
            print(f"[DEBUG] Aba RT não encontrada. Abas disponíveis: {wb_giga.sheetnames}")

        # Dados começam na linha 2 (índice 2)
        for r in giga_rows[2:]:
            tec = str(_get(r, C_TEC, "")).strip()
            if not tec or tec.lower() in ["none", "", "total geral", "total"]:
                continue
            # Combustível: tenta fórmula cached, senão cruza pela aba RT
            comb_raw = _get(r, C_COMB, None)
            if comb_raw is not None and comb_raw != "None":
                combustivel_val = round(_num(comb_raw), 2)
            else:
                combustivel_val = combustivel_map.get(tec.upper(), 0.0)
            indicadores_data.append({
                "tecnico":     tec,
                "ativacao":    int(_num(_get(r, C_ATV, 0))),
                "comodo":      int(_num(_get(r, C_COMODO, 0))),
                "endereco":    int(_num(_get(r, C_ENDER, 0))),
                "reparo":      int(_num(_get(r, C_REP, 0))),
                "upgrade":     int(_num(_get(r, C_UPG, 0))),
                "total":       int(_num(_get(r, C_TOT, 0))),
                "media_prod":  round(_num(_get(r, C_MPROD, 0)), 2),
                "pontuacao":   round(_num(_get(r, C_PONT, 0)), 1),
                "qnt_ifi_atv": int(_num(_get(r, C_QATV, 0))),
                "qnt_ifi_mde": int(_num(_get(r, C_QMDE, 0))),
                "qnt_irr":     int(_num(_get(r, C_QIRR, 0))),
                "nota_ifi_atv": _pct(_get(r, C_NATV, 0)),
                "nota_ifi_mde": _pct(_get(r, C_NMDE, 0)),
                "nota_irr":    _pct(_get(r, C_NIRR, 0)),
                "pont_pos":    round(_num(_get(r, C_POS, 0)), 1),
                "combustivel": combustivel_val,
            })
        print(f"[OK] {len(indicadores_data)} técnicos carregados do GIGA+")
        if indicadores_data:
            print(f"[DEBUG] 1º técnico: {indicadores_data[0]}")
    wb_giga.close()
else:
    print("[AVISO] giga_temp.xlsx n\u00e3o encontrado, aba Indicadores ficar\u00e1 vazia")

# KPIs globais (m\u00e9dias da equipe)
if indicadores_data:
    eq_media_prod = round(sum(x["media_prod"] for x in indicadores_data) / len(indicadores_data), 2)
    eq_pontuacao  = round(sum(x["pontuacao"]  for x in indicadores_data) / len(indicadores_data), 1)
    eq_nota_ifi   = round(sum(x["nota_ifi_mde"] for x in indicadores_data) / len(indicadores_data), 1)
    eq_nota_irr   = round(sum(x["nota_irr"]   for x in indicadores_data) / len(indicadores_data), 1)
    eq_total_os   = sum(x["total"] for x in indicadores_data)
    eq_combustivel = sum(x["combustivel"] for x in indicadores_data)
    top_prod = max(indicadores_data, key=lambda x: x["media_prod"])
else:
    eq_media_prod = eq_pontuacao = eq_nota_ifi = eq_nota_irr = 0
    eq_total_os = eq_combustivel = 0
    top_prod = {"tecnico": "-", "media_prod": 0}

def _nota_color(nota, inverso=False):
    """Verde = bom, Vermelho = ruim. inverso=True para IRR (0% = bom)"""
    if inverso:
        if nota <= 5:   return "var(--green)"
        if nota <= 15:  return "var(--yellow)"
        return "var(--red)"
    else:
        if nota >= 90:  return "var(--green)"
        if nota >= 70:  return "var(--yellow)"
        return "var(--red)"

def gen_indicadores_table(dados):
    rows_html = ""
    for d in dados:
        c_mp  = _nota_color(d["media_prod"] * 30, inverso=False) if d["media_prod"] > 0 else "var(--muted)"
        c_ifi = _nota_color(d["nota_ifi_mde"])
        c_irr = _nota_color(d["nota_irr"], inverso=True)
        comb_str = f"R$ {d['combustivel']:,.2f}" if d["combustivel"] > 0 else "-"
        rows_html += (
            f'<tr>'
            f'<td style="font-weight:600;color:#e8eaf6">{d["tecnico"]}</td>'
            f'<td style="text-align:center">{d["ativacao"]}</td>'
            f'<td style="text-align:center">{d["comodo"]}</td>'
            f'<td style="text-align:center">{d["endereco"]}</td>'
            f'<td style="text-align:center">{d["reparo"]}</td>'
            f'<td style="text-align:center">{d["upgrade"]}</td>'
            f'<td style="text-align:center;font-weight:700;color:var(--accent)">{d["total"]}</td>'
            f'<td style="text-align:center;font-weight:700;color:{c_mp}">{d["media_prod"]}</td>'
            f'<td style="text-align:center;font-weight:700;color:var(--accent2)">{d["pontuacao"]}</td>'
            f'<td style="text-align:center">{d["qnt_ifi_atv"]}</td>'
            f'<td style="text-align:center">{d["qnt_ifi_mde"]}</td>'
            f'<td style="text-align:center">{d["qnt_irr"]}</td>'
            f'<td style="text-align:center;font-weight:600;color:{_nota_color(d["nota_ifi_atv"])}">{d["nota_ifi_atv"]}%</td>'
            f'<td style="text-align:center;font-weight:600;color:{c_ifi}">{d["nota_ifi_mde"]}%</td>'
            f'<td style="text-align:center;font-weight:600;color:{c_irr}">{d["nota_irr"]}%</td>'
            f'<td style="text-align:center;font-weight:700;color:var(--yellow)">{d["pont_pos"]}</td>'
            f'<td style="text-align:right;color:#f472b6;font-weight:600">{comb_str}</td>'
            f'</tr>\n'
        )
    return rows_html

import sys
sys.path.insert(0, PASTA)
from indicadores_html_gen import gerar_indicadores_html
from whatsapp_agenda_gen import gerar_whatsapp_agenda_html
whatsapp_agenda_page_html = gerar_whatsapp_agenda_html()
# indicadores_page_html será gerado após bloco 2h (dados por técnico)

# --- 2g. INDICADOR GIGA (PAINEL DE METAS) ---
GIGA2_SHEET_ID = "121xl-QAkm8MNYM6hyQk2ueV2uBIybP_37O22zlj9HYg"
GIGA2_URL = f"https://docs.google.com/spreadsheets/d/{GIGA2_SHEET_ID}/export?format=xlsx"
GIGA2_FILE = os.path.join(PASTA, "premio_temp.xlsx")

try:
    import urllib.request as _ur2, shutil as _sh2
    _dl2 = os.path.join(PASTA, "premio_download.xlsx")
    print("[...] Baixando Indicador Giga...")
    _ur2.urlretrieve(GIGA2_URL, _dl2)
    _sh2.move(_dl2, GIGA2_FILE)
    print("[OK] premio_temp.xlsx atualizado!")
except Exception as _e2:
    print(f"[AVISO] Indicador Giga: {_e2}")

_IND_INV = {"IFI", "IRR", "IFIME", "OUTLIER"}
def _is_inv(nome):
    n = nome.upper()
    return any(k in n for k in _IND_INV)

import re as _re2
def _parse_meta(s):
    """'>85%' -> 85.0, '<2,7%' -> 2.7"""
    m = _re2.search(r'[\d,.]+', str(s).replace(',','.'))
    return float(m.group()) if m else 0.0

giga2_data = {}
if os.path.exists(GIGA2_FILE):
    _wb2 = openpyxl.load_workbook(GIGA2_FILE, read_only=True, data_only=True)
    if "PAINEL" in _wb2.sheetnames:
        _ws2 = _wb2["PAINEL"]
        _rows2 = list(_ws2.iter_rows(values_only=True))
        _mes2 = None
        _i2 = 0
        while _i2 < len(_rows2):
            _r2 = _rows2[_i2]
            _v0 = str(_r2[0]).strip() if _r2[0] else ""
            if not _v0:
                _i2 += 1; continue
            # Linha de MÊS
            if not _v0.startswith("C.") and _v0 not in ["INDICADORES", "TOTAL"] \
               and not any(_k in _v0.upper() for _k in ["EFICI","AGING","IFI","PRAZO","OUTL","IRR"]):
                _mes2 = _v0
                if _mes2 not in giga2_data: giga2_data[_mes2] = {}
                _i2 += 1; continue
            # Linha de CENTRAL
            if _v0.startswith("C.") and _mes2:
                _central2 = _v0.split(",")[0].strip()
                _datas2 = []
                _col2 = 4
                while _col2 < len(_r2):
                    _dt2 = _r2[_col2]
                    if _dt2 is None: break
                    if hasattr(_dt2, 'day'): _datas2.append(f"{_dt2.day:02d}/{_dt2.month:02d}")
                    _col2 += 2
                _inds2 = []
                for _j2 in range(_i2 + 2, min(_i2 + 11, len(_rows2))):
                    _ri2 = _rows2[_j2]
                    _ni2 = str(_ri2[0]).strip() if _ri2[0] else ""
                    if not _ni2 or "TOTAL" in _ni2.upper(): break
                    try: _pm2 = int(float(_ri2[3])) if _ri2[3] else 0
                    except: _pm2 = 0
                    _dias2 = []
                    for _di2 in range(len(_datas2)):
                        _cn2 = 4 + _di2 * 2; _cp2 = _cn2 + 1
                        _nota2 = float(_ri2[_cn2]) if _cn2 < len(_ri2) and _ri2[_cn2] is not None else None
                        _peso2 = float(_ri2[_cp2]) if _cp2 < len(_ri2) and _ri2[_cp2] is not None else 0.0
                        _dias2.append({"n": round(_nota2 * 100, 1) if _nota2 is not None else None, "p": int(_peso2)})
                    _m100s = str(_ri2[1]).strip() if _ri2[1] else ""
                    _m80s = str(_ri2[2]).strip() if _ri2[2] else ""
                    _inds2.append({"nome": _ni2, "m100": _m100s, "m80": _m80s,
                                   "m100v": _parse_meta(_m100s), "m80v": _parse_meta(_m80s),
                                   "pmax": _pm2, "inv": _is_inv(_ni2), "dias": _dias2})
                _tidx2 = _i2 + 2 + len(_inds2)
                _tots2 = []
                if _tidx2 < len(_rows2):
                    _rt2 = _rows2[_tidx2]
                    for _di2 in range(len(_datas2)):
                        _ct2 = 4 + _di2 * 2
                        _vt2 = _rt2[_ct2] if _ct2 < len(_rt2) else None
                        _tots2.append(int(_vt2) if _vt2 is not None else 0)
                giga2_data[_mes2][_central2] = {"datas": _datas2, "inds": _inds2, "tots": _tots2}
                _i2 = _tidx2 + 1; continue
            _i2 += 1
        _wb2.close()
    print(f"[OK] Indicador Giga: {list(giga2_data.keys())} meses")
else:
    print("[AVISO] premio_temp.xlsx nao encontrado")

# --- 2h. INDICADORES POR TÉCNICO (ABAS DETALHADAS) ---
import re

def _norm_tec_name(nome):
    if not nome: return ""
    nome = str(nome).strip()
    nome = re.sub(r'\s*-\s*SVOBODA.*$', '', nome, flags=re.IGNORECASE)
    nome = nome.strip()
    if nome == nome.upper() or nome == nome.lower():
        nome = nome.title()
    return nome

def _detect_month_suffix(sheetnames):
    meses_ord = ["JAN","FEV","MAR","ABR","MAI","JUN","JUL","AGO","SET","OUT","NOV","DEZ"]
    found = set()
    for sn in sheetnames:
        for m in meses_ord:
            if sn.upper().strip().endswith(m):
                found.add(m)
    for m in reversed(meses_ord):
        if m in found:
            return m
    return "ABR"

tec_altas_data = []
tec_reparo_data = []
tec_irr_data = []
tec_ifi_data = []
tec_efetiv_data = []
raw_irr_list = []
raw_ifi_list = []
# Contadores para cálculo IFI% e IFIME%
_tec_alta_count = {}  # {nome_normalizado: qtd_ativações}
_tec_me_count = {}    # {nome_normalizado: qtd_mudancas_endereco}
_tec_cluster = {}     # {nome_normalizado: cluster}

if os.path.exists(GIGA2_FILE):
    _wb3 = openpyxl.load_workbook(GIGA2_FILE, read_only=True, data_only=True)
    _mes_suffix = _detect_month_suffix(_wb3.sheetnames)
    print(f"[DEBUG] Mês detectado para indicadores: {_mes_suffix}")

    # ── ALTAS ──
    _aba_altas = next((n for n in _wb3.sheetnames if n.upper().strip().startswith("ALTAS") and n.upper().strip().endswith(_mes_suffix) and "CORRETO" not in n.upper()), None)
    if _aba_altas:
        _ws_a = _wb3[_aba_altas]
        _rows_a = list(_ws_a.iter_rows(values_only=True))
        _agg_a = {}
        for r in _rows_a[1:]:
            tec = _norm_tec_name(r[11] if len(r) > 11 else "")
            if not tec: continue
            aging_h = 0
            try: aging_h = float(r[14]) if len(r) > 14 and r[14] else 0
            except: pass
            aging_cat = str(r[15]).strip() if len(r) > 15 and r[15] else ""
            outlier = str(r[17]).strip().upper() if len(r) > 17 and r[17] else ""
            if tec not in _agg_a:
                _agg_a[tec] = {"qtd": 0, "aging_sum": 0, "dentro_48h": 0, "outlier": 0}
            _agg_a[tec]["qtd"] += 1
            _agg_a[tec]["aging_sum"] += aging_h
            if "DENTRO" in aging_cat.upper():
                _agg_a[tec]["dentro_48h"] += 1
            if outlier in ["VERDADEIRO", "TRUE", "1", "SIM"]:
                _agg_a[tec]["outlier"] += 1
            # Contadores para IFI% e IFIME%
            _tipo_al = str(r[0]).strip().upper() if r[0] else ""
            _cluster_al = str(r[5]).strip() if len(r) > 5 and r[5] else ""
            if _cluster_al and tec not in _tec_cluster:
                _tec_cluster[tec] = _cluster_al
            if _tipo_al == "ALTA":
                _tec_alta_count[tec] = _tec_alta_count.get(tec, 0) + 1
            elif _tipo_al == "ME":
                _tec_me_count[tec] = _tec_me_count.get(tec, 0) + 1
        for tec, d in sorted(_agg_a.items(), key=lambda x: x[1]["qtd"], reverse=True):
            tec_altas_data.append({
                "tecnico": tec, "qtd": d["qtd"],
                "aging_medio": round(d["aging_sum"] / d["qtd"], 1) if d["qtd"] > 0 else 0,
                "pct_48h": round(d["dentro_48h"] / d["qtd"] * 100, 1) if d["qtd"] > 0 else 0,
                "pct_outlier": round(d["outlier"] / d["qtd"] * 100, 1) if d["qtd"] > 0 else 0,
            })
        print(f"[OK] Altas: {len(tec_altas_data)} técnicos de {len(_rows_a)-1} registros")

    # ── REPARO ──
    _aba_rep = next((n for n in _wb3.sheetnames if n.upper().strip().startswith("REPARO") and n.upper().strip().endswith(_mes_suffix) and "CORRETO" not in n.upper()), None)
    if _aba_rep:
        _ws_r3 = _wb3[_aba_rep]
        _rows_r3 = list(_ws_r3.iter_rows(values_only=True))
        _agg_r3 = {}
        for r in _rows_r3[1:]:
            tec = _norm_tec_name(r[11] if len(r) > 11 else "")
            if not tec: continue
            aging_h = 0
            try: aging_h = float(r[13]) if len(r) > 13 and r[13] else 0
            except: pass
            aging_cat = str(r[14]).strip() if len(r) > 14 and r[14] else ""
            if tec not in _agg_r3:
                _agg_r3[tec] = {"qtd": 0, "aging_sum": 0, "dentro_24h": 0}
            _agg_r3[tec]["qtd"] += 1
            _agg_r3[tec]["aging_sum"] += aging_h
            if "18" in aging_cat or "24" in aging_cat:
                _agg_r3[tec]["dentro_24h"] += 1
        for tec, d in sorted(_agg_r3.items(), key=lambda x: x[1]["qtd"], reverse=True):
            tec_reparo_data.append({
                "tecnico": tec, "qtd": d["qtd"],
                "aging_medio": round(d["aging_sum"] / d["qtd"], 1) if d["qtd"] > 0 else 0,
                "pct_24h": round(d["dentro_24h"] / d["qtd"] * 100, 1) if d["qtd"] > 0 else 0,
            })
        print(f"[OK] Reparo: {len(tec_reparo_data)} técnicos de {len(_rows_r3)-1} registros")
        # Mapear contagem de reparos por técnico para cálculo IRR%
        _tec_reparo_count = {d["tecnico"]: d["qtd"] for d in tec_reparo_data}

    # ── IRR ──
    _METAS_IRR = {"C.15": (4.2, 4.4, 20), "C.16": (6.2, 6.5, 20), "C.17": (6.2, 6.5, 20), "C.18": (4.2, 5.2, 20)}
    _aba_irr3 = next((n for n in _wb3.sheetnames if n.upper().strip().startswith("IRR") and n.upper().strip().endswith(_mes_suffix) and "CORRETO" not in n.upper()), None)
    if _aba_irr3:
        _ws_i3 = _wb3[_aba_irr3]
        _rows_i3 = list(_ws_i3.iter_rows(values_only=True))
        _agg_irr = {}
        for r in _rows_i3[1:]:
            tec = _norm_tec_name(r[16] if len(r) > 16 else "")
            if not tec: continue
            cidade = str(r[4]).strip() if len(r) > 4 and r[4] else ""
            motivo = str(r[11]).strip() if len(r) > 11 and r[11] else ""
            os_cod = str(r[10]).strip() if len(r) > 10 and r[10] else ""
            data_str = ""
            if len(r) > 18 and r[18]:
                try: data_str = r[18].strftime("%d/%m/%Y") if hasattr(r[18], 'strftime') else str(r[18])[:10]
                except: data_str = str(r[18])[:10]
            raw_irr_list.append({"tecnico": tec, "cidade": cidade, "motivo": motivo, "os": os_cod, "data": data_str})
            if tec not in _agg_irr: _agg_irr[tec] = 0
            _agg_irr[tec] += 1
        # Gerar tec_irr_data com % baseado em reparos
        _all_tec_irr = set(list(_agg_irr.keys()) + list(_tec_reparo_count.keys()))
        for tec in _all_tec_irr:
            qtd_irr = _agg_irr.get(tec, 0)
            reparos = _tec_reparo_count.get(tec, 0)
            cluster = _tec_cluster.get(tec, "")
            pct = round(qtd_irr / reparos * 100, 2) if reparos > 0 else 0.0
            m100, m80, peso = _METAS_IRR.get(cluster, (6.2, 6.5, 20))
            nota = peso if pct < m100 else round(peso * 0.8) if pct < m80 else 0
            tec_irr_data.append({"tecnico": tec, "qtd": qtd_irr, "reparos": reparos, "cluster": cluster, "pct_irr": pct, "meta100": m100, "nota": nota})
        tec_irr_data.sort(key=lambda x: x["pct_irr"], reverse=True)
        print(f"[OK] IRR: {len(tec_irr_data)} técnicos, {sum(d['qtd'] for d in tec_irr_data)} casos")

    # ── IFI + IFIME ──
    _METAS_IFI = {"C.15": (2.7, 3.3, 15), "C.16": (2.7, 3.3, 15), "C.17": (2.7, 3.3, 15), "C.18": (2.7, 3.3, 15)}
    _METAS_IFIME = {"C.15": (3.25, 3.9, 10), "C.16": (3.25, 3.9, 10), "C.17": (3.25, 3.9, 10), "C.18": (3.25, 3.9, 10)}
    raw_ifime_list = []
    tec_ifime_data = []
    _aba_ifi3 = next((n for n in _wb3.sheetnames if n.upper().strip().startswith("IFI") and n.upper().strip().endswith(_mes_suffix) and "CORRETO" not in n.upper()), None)
    if _aba_ifi3:
        _ws_f3 = _wb3[_aba_ifi3]
        _rows_f3 = list(_ws_f3.iter_rows(values_only=True))
        _agg_ifi_pure = {}  # IFI puro (sem IFME)
        _agg_ifime = {}     # IFIME
        for r in _rows_f3[1:]:
            tec = _norm_tec_name(r[14] if len(r) > 14 else "")
            if not tec: continue
            ifi_flag = str(r[1]).strip().upper() if len(r) > 1 and r[1] else ""
            ifme_flag = str(r[2]).strip().upper() if len(r) > 2 and r[2] else ""
            cidade = str(r[5]).strip() if len(r) > 5 and r[5] else ""
            motivo = str(r[12]).strip() if len(r) > 12 and r[12] else ""
            os_cod = str(r[11]).strip() if len(r) > 11 and r[11] else ""
            data_str = ""
            if len(r) > 16 and r[16]:
                try: data_str = r[16].strftime("%d/%m/%Y") if hasattr(r[16], 'strftime') else str(r[16])[:10]
                except: data_str = str(r[16])[:10]
            is_ifi = ifi_flag in ["VERDADEIRO", "TRUE", "1", "SIM"]
            is_ifme = ifme_flag in ["VERDADEIRO", "TRUE", "1", "SIM"]
            if is_ifme:
                raw_ifime_list.append({"tecnico": tec, "cidade": cidade, "motivo": motivo, "os": os_cod, "data": data_str})
                if tec not in _agg_ifime: _agg_ifime[tec] = 0
                _agg_ifime[tec] += 1
            elif is_ifi:
                raw_ifi_list.append({"tecnico": tec, "cidade": cidade, "motivo": motivo, "os": os_cod, "data": data_str, "ifme": False})
                if tec not in _agg_ifi_pure: _agg_ifi_pure[tec] = 0
                _agg_ifi_pure[tec] += 1
        # Gerar tec_ifi_data com % baseado em ativações (ALTA)
        _all_tec_ifi = set(list(_agg_ifi_pure.keys()) + list(_tec_alta_count.keys()))
        for tec in _all_tec_ifi:
            qtd_ifi = _agg_ifi_pure.get(tec, 0)
            altas = _tec_alta_count.get(tec, 0)
            cluster = _tec_cluster.get(tec, "")
            pct = round(qtd_ifi / altas * 100, 2) if altas > 0 else 0.0
            m100, m80, peso = _METAS_IFI.get(cluster, (2.7, 3.3, 15))
            nota = peso if pct < m100 else round(peso * 0.8) if pct < m80 else 0
            tec_ifi_data.append({"tecnico": tec, "qtd": qtd_ifi, "altas": altas, "cluster": cluster, "pct_ifi": pct, "meta100": m100, "nota": nota, "qtd_ifme": 0})
        tec_ifi_data.sort(key=lambda x: x["pct_ifi"], reverse=True)
        # Gerar tec_ifime_data com % baseado em ME
        _all_tec_ifime = set(list(_agg_ifime.keys()) + list(_tec_me_count.keys()))
        for tec in _all_tec_ifime:
            qtd_ifme = _agg_ifime.get(tec, 0)
            mes = _tec_me_count.get(tec, 0)
            cluster = _tec_cluster.get(tec, "")
            pct = round(qtd_ifme / mes * 100, 2) if mes > 0 else 0.0
            m100, m80, peso = _METAS_IFIME.get(cluster, (3.25, 3.9, 10))
            nota = peso if pct < m100 else round(peso * 0.8) if pct < m80 else 0
            tec_ifime_data.append({"tecnico": tec, "qtd": qtd_ifme, "mes_total": mes, "cluster": cluster, "pct_ifime": pct, "meta100": m100, "nota": nota})
        tec_ifime_data.sort(key=lambda x: x["pct_ifime"], reverse=True)
        print(f"[OK] IFI: {len(tec_ifi_data)} técnicos, {sum(d['qtd'] for d in tec_ifi_data)} casos")
        print(f"[OK] IFIME: {len(tec_ifime_data)} técnicos, {sum(d['qtd'] for d in tec_ifime_data)} casos")

    # ── EFETIVIDADE ──
    _aba_efe3 = next((n for n in _wb3.sheetnames if n.upper().strip().startswith("EFETIVIDADE") and n.upper().strip().endswith(_mes_suffix)), None)
    if _aba_efe3:
        _ws_e3 = _wb3[_aba_efe3]
        _rows_e3 = list(_ws_e3.iter_rows(values_only=True))
        _agg_e3 = {}
        for r in _rows_e3[1:]:
            tec = _norm_tec_name(r[8] if len(r) > 8 else "")
            if not tec: continue
            conclusao = str(r[7]).strip() if len(r) > 7 and r[7] else ""
            if tec not in _agg_e3: _agg_e3[tec] = {"qtd": 0, "realizado": 0}
            _agg_e3[tec]["qtd"] += 1
            if "realizado" in conclusao.lower(): _agg_e3[tec]["realizado"] += 1
        for tec, d in sorted(_agg_e3.items(), key=lambda x: x[1]["qtd"], reverse=True):
            tec_efetiv_data.append({
                "tecnico": tec, "qtd": d["qtd"], "realizado": d["realizado"],
                "pct_realizado": round(d["realizado"] / d["qtd"] * 100, 1) if d["qtd"] > 0 else 0,
            })
        print(f"[OK] Efetividade: {len(tec_efetiv_data)} técnicos de {len(_rows_e3)-1} registros")

    _wb3.close()
else:
    print("[AVISO] premio_temp.xlsx não encontrado para indicadores por técnico")

_total_altas = sum(d["qtd"] for d in tec_altas_data)
_total_reparos_ind = sum(d["qtd"] for d in tec_reparo_data)
_total_irr_ind = sum(d["qtd"] for d in tec_irr_data)
_total_ifi_ind = sum(d["qtd"] for d in tec_ifi_data)
_total_efetiv = sum(d["qtd"] for d in tec_efetiv_data)
_avg_aging_altas = round(sum(d["aging_medio"]*d["qtd"] for d in tec_altas_data) / _total_altas, 1) if _total_altas > 0 else 0
_avg_aging_reparo = round(sum(d["aging_medio"]*d["qtd"] for d in tec_reparo_data) / _total_reparos_ind, 1) if _total_reparos_ind > 0 else 0

# ── SCORECARD COMPARATIVO ──
_idx_altas = {d["tecnico"]: d for d in tec_altas_data}
_idx_reparo = {d["tecnico"]: d for d in tec_reparo_data}
_idx_irr = {d["tecnico"]: d for d in tec_irr_data}
_idx_ifi = {d["tecnico"]: d for d in tec_ifi_data}
_idx_ifime = {d["tecnico"]: d for d in tec_ifime_data}
_idx_efetiv = {d.get("Recurso", d.get("key","")): d for d in (list_tec if list_tec else [])}
_METAS_AGING = {"C.15": (85, 80, 15), "C.16": (85, 80, 15), "C.17": (85, 80, 15), "C.18": (85, 80, 15)}
_METAS_REP24 = {"C.15": (85, 80, 15), "C.16": (85, 80, 15), "C.17": (85, 80, 15), "C.18": (85, 80, 15)}
_METAS_OUTLIER = {"C.15": (3, 3.8, 5), "C.16": (3, 3.8, 5), "C.17": (3, 3.8, 5), "C.18": (3, 3.8, 5)}
_METAS_EFETIV = {"C.15": (85, 80, 20), "C.16": (85, 80, 20), "C.17": (85, 80, 20), "C.18": (85, 80, 20)}
_all_tecs = set()
for lst in [tec_altas_data, tec_reparo_data, tec_irr_data, tec_ifi_data, tec_ifime_data]:
    for d in lst:
        _all_tecs.add(d["tecnico"])
tec_scorecard = []
for tec in _all_tecs:
    cluster = _tec_cluster.get(tec, "")
    # Aging 48h
    a = _idx_altas.get(tec, {})
    pct_48 = a.get("pct_48h", 0)
    m100, m80, p = _METAS_AGING.get(cluster, (85, 80, 15))
    n_aging = p if pct_48 >= m100 else round(p*0.8) if pct_48 >= m80 else 0
    # Reparo 24h
    rp = _idx_reparo.get(tec, {})
    pct_24 = rp.get("pct_24h", 0)
    m100, m80, p = _METAS_REP24.get(cluster, (85, 80, 15))
    n_rep24 = p if pct_24 >= m100 else round(p*0.8) if pct_24 >= m80 else 0
    # Outlier (0% outlier = bom = nota máxima)
    pct_out = a.get("pct_outlier", 0)
    m100, m80, p = _METAS_OUTLIER.get(cluster, (3, 3.8, 5))
    n_outlier = p if pct_out <= m100 else round(p*0.8) if pct_out <= m80 else 0
    # IFI (0 casos = 0% = nota máxima, técnico sem ocorrência)
    fi = _idx_ifi.get(tec, {})
    pct_ifi = fi.get("pct_ifi", 0)
    if tec in _idx_ifi:
        n_ifi = fi.get("nota", 0)
    else:
        _m100_ifi, _m80_ifi, _p_ifi = _METAS_IFI.get(cluster, (2.7, 3.3, 15))
        n_ifi = _p_ifi  # 0% está dentro da meta = nota máxima
    # IRR (0 casos = 0% = nota máxima)
    ir = _idx_irr.get(tec, {})
    pct_irr = ir.get("pct_irr", 0)
    if tec in _idx_irr:
        n_irr = ir.get("nota", 0)
    else:
        _m100_irr, _m80_irr, _p_irr = _METAS_IRR.get(cluster, (6.2, 6.5, 20))
        n_irr = _p_irr  # 0% = nota máxima
    # IFIME (0 casos = 0% = nota máxima)
    fm = _idx_ifime.get(tec, {})
    pct_ifime = fm.get("pct_ifime", 0)
    if tec in _idx_ifime:
        n_ifime = fm.get("nota", 0)
    else:
        _m100_ifm, _m80_ifm, _p_ifm = _METAS_IFIME.get(cluster, (3.25, 3.9, 10))
        n_ifime = _p_ifm  # 0% = nota máxima
    # Efetividade
    ef = _idx_efetiv.get(tec, {})
    taxa_ef = ef.get("taxa", 0)
    m100, m80, p = _METAS_EFETIV.get(cluster, (85, 80, 20))
    n_efetiv = p if taxa_ef >= m100 else round(p*0.8) if taxa_ef >= m80 else 0
    total = n_aging + n_rep24 + n_outlier + n_ifi + n_irr + n_ifime + n_efetiv
    tec_scorecard.append({
        "tecnico": tec, "cluster": cluster,
        "pct_48h": pct_48, "n_aging": n_aging,
        "pct_24h": pct_24, "n_rep24": n_rep24,
        "pct_outlier": pct_out, "n_outlier": n_outlier,
        "pct_ifi": pct_ifi, "n_ifi": n_ifi,
        "pct_irr": pct_irr, "n_irr": n_irr,
        "pct_ifime": pct_ifime, "n_ifime": n_ifime,
        "taxa_efetiv": taxa_ef, "n_efetiv": n_efetiv,
        "total": total
    })
tec_scorecard.sort(key=lambda x: x["total"], reverse=True)
print(f"[OK] Scorecard: {len(tec_scorecard)} técnicos, média={round(sum(d['total'] for d in tec_scorecard)/max(len(tec_scorecard),1),1)}")

# Gera o HTML completo da aba Indicadores (com sub-abas)
indicadores_page_html = gerar_indicadores_html(
    tec_altas_data, tec_reparo_data, tec_irr_data, tec_ifi_data, tec_efetiv_data,
    raw_irr_list, raw_ifi_list, indicadores_data, gen_indicadores_table,
    eq_total_os, eq_media_prod, eq_pontuacao, eq_nota_ifi,
    eq_nota_irr, eq_combustivel, _nota_color, list_tec,
    tec_ifime_data, raw_ifime_list, tec_scorecard
)

_MORD = ["DEZEMBRO", "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO"]
_mdisp = [m for m in _MORD if m in giga2_data]
_mdef = _mdisp[-1] if _mdisp else ""
_giga2_json = json.dumps(giga2_data, ensure_ascii=False)
_month_tabs_html = "".join(
    f'<div class="ig-month-tab{"  active" if m == _mdef else ""}" onclick="igSetMes(this,\'{m}\')">'
    f'{m.capitalize()}</div>'
    for m in _mdisp
)

indicador_giga_page_html = f"""
<!-- ==================== INDICADOR GIGA ==================== -->
<style>
  /* ── INDICADOR GIGA ─────────────────────────────────────── */
  .ig-month-tabs {{display:flex;gap:8px;margin-bottom:20px;flex-wrap:wrap;align-items:center;}}
  .ig-month-tab {{background:#0d1520;border:1px solid #1c2237;border-radius:10px;padding:8px 20px;font-size:13px;color:#64748b;cursor:pointer;font-family:var(--font-head);font-weight:600;transition:all 0.2s;letter-spacing:0.3px;}}
  .ig-month-tab:hover {{border-color:#7c3aed44;color:#a78bfa;}}
  .ig-month-tab.active {{background:linear-gradient(135deg,#7c3aed,#6d28d9);border-color:transparent;color:#fff;box-shadow:0 4px 18px rgba(124,58,237,0.35);transform:translateY(-1px);}}
  .ig-sector-tabs {{display:flex;gap:10px;margin-bottom:24px;align-items:center;}}
  .ig-sector-tab {{background:#0d1520;border:2px solid #1c2237;border-radius:12px;padding:10px 26px;font-size:15px;color:#64748b;cursor:pointer;font-family:var(--font-head);font-weight:800;transition:all 0.2s;letter-spacing:0.5px;}}
  .ig-sector-tab:hover {{border-color:#22d3a044;color:#6ee7b7;}}
  .ig-sector-tab.active {{background:linear-gradient(135deg,#22d3a0,#059669);border-color:transparent;color:#022c22;box-shadow:0 4px 20px rgba(34,211,160,0.3);transform:translateY(-1px);}}
  /* KPI row */
  .ig-kpi-row {{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:28px;}}
  .ig-kpi {{background:#0d1520;border:1px solid #1c2237;border-radius:16px;padding:20px 22px;display:flex;flex-direction:column;gap:4px;transition:box-shadow 0.2s;}}
  .ig-kpi:hover {{box-shadow:0 0 0 1px #7c3aed33,0 8px 30px rgba(0,0,0,0.3);}}
  .ig-kpi-lbl {{font-size:10px;color:#475569;text-transform:uppercase;letter-spacing:1.2px;font-weight:700;font-family:var(--font-head);}}
  .ig-kpi-val {{font-size:38px;font-weight:900;font-family:var(--font-head);line-height:1.1;letter-spacing:-1px;}}
  .ig-kpi-sub {{font-size:12px;color:#334155;margin-top:2px;}}
  /* Gauge grid – 4 cols, last row centered via auto-placement */
  .ig-gauges-grid {{display:grid;grid-template-columns:repeat(4,1fr);gap:16px;margin-bottom:24px;}}
  .ig-gauge-card {{background:#0d1520;border:1px solid #1c2237;border-radius:20px;padding:28px 18px 20px;display:flex;flex-direction:column;align-items:center;gap:10px;transition:all 0.25s;position:relative;overflow:hidden;}}
  .ig-gauge-card:hover {{transform:translateY(-4px);box-shadow:0 16px 48px rgba(0,0,0,0.45);}}
  .ig-gauge-card::after {{content:'';position:absolute;top:0;left:0;right:0;height:2px;border-radius:20px 20px 0 0;}}
  .ig-gauge-card.ig-green {{border-color:#22d3a022;background:linear-gradient(160deg,#0d1520,#091a13);}}
  .ig-gauge-card.ig-green::after {{background:linear-gradient(90deg,#22d3a0,#059669);}}
  .ig-gauge-card.ig-yellow {{border-color:#fbbf2422;background:linear-gradient(160deg,#0d1520,#191208);}}
  .ig-gauge-card.ig-yellow::after {{background:linear-gradient(90deg,#fbbf24,#d97706);}}
  .ig-gauge-card.ig-red {{border-color:#ff4d6d22;background:linear-gradient(160deg,#0d1520,#190810);}}
  .ig-gauge-card.ig-red::after {{background:linear-gradient(90deg,#ff4d6d,#e11d48);}}
  .ig-gauge-card.ig-gray {{border-color:#1c2237;}}
  .ig-gauge-card.ig-gray::after {{background:#1c2237;}}
  .ig-gauge-name {{font-size:13px;color:#cbd5e1;text-align:center;font-family:var(--font-head);font-weight:700;line-height:1.35;max-width:180px;}}
  .ig-gauge-meta {{font-size:11px;color:#334155;text-align:center;font-family:var(--font-mono);}}
  .ig-gauge-pts {{font-size:13px;font-weight:700;font-family:var(--font-mono);padding:5px 14px;border-radius:20px;margin-top:2px;}}
  .ig-gauge-pts.ig-green {{background:rgba(34,211,160,0.1);color:#22d3a0;border:1px solid rgba(34,211,160,0.25);}}
  .ig-gauge-pts.ig-yellow {{background:rgba(251,191,36,0.1);color:#fbbf24;border:1px solid rgba(251,191,36,0.25);}}
  .ig-gauge-pts.ig-red {{background:rgba(255,77,109,0.1);color:#ff4d6d;border:1px solid rgba(255,77,109,0.25);}}
  .ig-gauge-pts.ig-gray {{background:rgba(100,116,139,0.08);color:#64748b;border:1px solid rgba(100,116,139,0.15);}}
  /* Evolução */
  .ig-evo {{background:#0d1520;border:1px solid #1c2237;border-radius:16px;padding:26px;}}
  .ig-evo-title {{font-size:14px;font-weight:700;color:#e2e8f0;font-family:var(--font-head);margin-bottom:22px;display:flex;align-items:center;gap:8px;}}
  .ig-evo-wrap {{display:flex;align-items:flex-end;gap:8px;height:140px;}}
  .ig-evo-col {{display:flex;flex-direction:column;align-items:center;gap:6px;flex:1;cursor:pointer;}}
  .ig-evo-bar {{width:100%;border-radius:8px 8px 0 0;min-height:4px;transition:height 0.4s cubic-bezier(.4,0,.2,1),opacity 0.3s;}}
  .ig-evo-col:hover .ig-evo-bar {{opacity:1!important;filter:brightness(1.2);}}
  .ig-evo-num {{font-size:13px;font-weight:800;font-family:var(--font-mono);}}
  .ig-evo-dt {{font-size:11px;color:#475569;font-family:var(--font-mono);}}
  /* Day selector */
  .ig-day-selector {{display:flex;gap:6px;margin-bottom:22px;flex-wrap:wrap;align-items:center;}}
  .ig-day-btn {{background:#0d1520;border:1px solid #1c2237;border-radius:8px;padding:7px 14px;font-size:12px;color:#64748b;cursor:pointer;font-family:var(--font-mono);font-weight:600;transition:all 0.15s;}}
  .ig-day-btn:hover {{border-color:#7c3aed66;color:#a78bfa;background:#0f1628;}}
  .ig-day-btn.active {{background:rgba(124,58,237,0.15);border-color:#7c3aed;color:#c4b5fd;font-weight:800;}}
</style>
<div class="page" id="indicador-giga">
  <div class="ig-month-tabs">{_month_tabs_html}</div>
  <div class="ig-sector-tabs" id="ig-sectors"></div>
  <div class="ig-day-selector" id="ig-days"></div>
  <div class="ig-kpi-row" id="ig-kpis"></div>
  <div class="ig-gauges-grid" id="ig-gauges"></div>
  <div class="ig-evo">
    <div class="ig-evo-title">📈 Evolução Total de Pontos por Dia</div>
    <div class="ig-evo-wrap" id="ig-evo"></div>
  </div>
  <script>
  var _GIGA2 = {_giga2_json};
  var _igMes = {json.dumps(_mdef)};
  var _igC = null;
  var _igDay = null;

  function igSetMes(el, mes) {{
    document.querySelectorAll('.ig-month-tab').forEach(function(t){{t.classList.remove('active');}});
    el.classList.add('active');
    _igMes = mes; _igC = null; _igDay = null;
    igRenderSectors();
  }}
  function igSetC(el, c) {{
    document.querySelectorAll('.ig-sector-tab').forEach(function(t){{t.classList.remove('active');}});
    el.classList.add('active');
    _igC = c; _igDay = null;
    igRenderDays();
    igRender();
  }}
  function igSetDay(el, idx) {{
    document.querySelectorAll('.ig-day-btn').forEach(function(t){{t.classList.remove('active');}});
    el.classList.add('active');
    _igDay = idx;
    igRender();
  }}
  function igRenderSectors() {{
    var div = document.getElementById('ig-sectors'); div.innerHTML = '';
    var data = _GIGA2[_igMes]; if (!data) return;
    var cs = Object.keys(data);
    if (!_igC || !data[_igC]) _igC = cs[0];
    cs.forEach(function(c) {{
      var btn = document.createElement('div');
      btn.className = 'ig-sector-tab' + (c === _igC ? ' active' : '');
      btn.textContent = c;
      btn.onclick = function(){{ igSetC(this, c); }};
      div.appendChild(btn);
    }});
    igRenderDays(); igRender();
  }}
  function igRenderDays() {{
    var div = document.getElementById('ig-days'); div.innerHTML = '';
    var bloco = (_GIGA2[_igMes] || {{}})[_igC]; if (!bloco) return;
    bloco.datas.forEach(function(dt, idx) {{
      if (bloco.tots[idx] === 0 && idx > 0) return;
      var btn = document.createElement('div');
      btn.className = 'ig-day-btn' + (idx === _igDay ? ' active' : '');
      btn.textContent = dt;
      btn.onclick = function(){{ igSetDay(this, idx); }};
      div.appendChild(btn);
    }});
    // Default to last filled day
    if (_igDay === null) {{
      var lastFilled = 0;
      bloco.datas.forEach(function(_, idx){{ if(bloco.tots[idx] > 0) lastFilled = idx; }});
      _igDay = lastFilled;
      var btns = div.querySelectorAll('.ig-day-btn');
      var activable = [];
      bloco.datas.forEach(function(dt, idx){{ if(bloco.tots[idx] > 0 || idx === 0) activable.push(idx); }});
      var pos = activable.indexOf(_igDay);
      if (pos >= 0 && btns[pos]) btns[pos].classList.add('active');
    }}
  }}
  function igGaugeSvg(notaPct, color, label, m80v, m100v, isInv, rawVal) {{
    var pct = notaPct !== null ? Math.min(Math.max(notaPct / 100, 0), 1) : 0;
    var r = 68, cx = 90, cy = 94, sw = 14;
    var arcD = 'M '+(cx-r)+' '+cy+' A '+r+' '+r+' 0 0 1 '+(cx+r)+' '+cy;
    var svg = '';
    // Background track
    svg += '<path d="'+arcD+'" fill="none" stroke="#1e293b" stroke-width="'+sw+'" pathLength="100" stroke-linecap="round"/>';
    // Value arc
    if (pct > 0) {{
      var dash = (pct * 100).toFixed(1);
      svg += '<path d="'+arcD+'" fill="none" stroke="'+color+'" stroke-width="'+(sw+8)+'" pathLength="100" stroke-dasharray="'+dash+' 100" stroke-linecap="round" opacity="0.1"/>';
      svg += '<path d="'+arcD+'" fill="none" stroke="'+color+'" stroke-width="'+sw+'" pathLength="100" stroke-dasharray="'+dash+' 100" stroke-linecap="round"/>';
    }}
    // Meta markers
    function drawMark(mFrac, mColor, mLabel) {{
      var a = Math.PI * (1 - mFrac);
      var ca = Math.cos(a), sa = Math.sin(a);
      var ri = r - sw/2 - 4, ro = r + sw/2 + 4;
      svg += '<line x1="'+(cx+ri*ca).toFixed(1)+'" y1="'+(cy-ri*sa).toFixed(1)+'" x2="'+(cx+ro*ca).toFixed(1)+'" y2="'+(cy-ro*sa).toFixed(1)+'" stroke="'+mColor+'" stroke-width="3" stroke-linecap="round"/>';
      var rl = r - 22;
      svg += '<text x="'+(cx+rl*ca).toFixed(1)+'" y="'+(cy-rl*sa+3).toFixed(1)+'" text-anchor="middle" fill="'+mColor+'" font-size="9" font-weight="700" font-family="Outfit,sans-serif">'+mLabel+'</text>';
    }}
    if (m80v > 0 && m100v > 0) {{
      var mp80, mp100;
      if (isInv) {{
        var sc = Math.max(m80v * 2, (rawVal || 0) * 1.3);
        mp80 = Math.min(0.95, (sc - m80v) / sc);
        mp100 = Math.min(0.95, (sc - m100v) / sc);
      }} else {{
        mp80 = m80v / 100;
        mp100 = m100v / 100;
      }}
      drawMark(mp100, '#22d3a0', m100v+'%');
    }}
    var valTxt = label !== null ? label : '--';
    svg += '<text x="'+cx+'" y="'+(cy-18)+'" text-anchor="middle" fill="'+color+'" font-size="26" font-weight="900" font-family="Outfit,sans-serif" letter-spacing="-0.8">'+valTxt+'</text>';
    return '<svg viewBox="0 0 180 94" width="180" height="94" style="display:block;margin:0 auto;overflow:hidden;">' + svg + '</svg>';
  }}
  function igColorCls(p, pmax) {{
    if (!pmax) return {{c:'#475569',cls:'ig-gray'}};
    if (p >= pmax) return {{c:'#22d3a0',cls:'ig-green'}};
    if (p > 0) return {{c:'#fbbf24',cls:'ig-yellow'}};
    return {{c:'#ff4d6d',cls:'ig-red'}};
  }}
  function igRender() {{
    var bloco = (_GIGA2[_igMes] || {{}})[_igC]; if (!bloco) return;
    var datas = bloco.datas, inds = bloco.inds, tots = bloco.tots;
    var di = _igDay !== null ? _igDay : 0;

    // KPI cards
    var kDiv = document.getElementById('ig-kpis'); kDiv.innerHTML = '';
    var tot = tots[di] || 0;
    var maxPts = inds.reduce(function(s,x){{return s + x.pmax;}}, 0);
    var tColor = tot >= 85 ? '#22d3a0' : tot >= 60 ? '#fbbf24' : '#ff4d6d';
    var tBorder = tot >= 85 ? 'rgba(34,211,160,.3)' : tot >= 60 ? 'rgba(251,191,36,.3)' : 'rgba(255,77,109,.3)';
    var melhor = inds.reduce(function(b,ind) {{
      var d = ind.dias[di];
      if (!d) return b;
      if (!b || d.p > b.p) return {{nome: ind.nome, p: d.p, pmax: ind.pmax}};
      return b;
    }}, null);
    var alertas = inds.filter(function(ind){{ var d = ind.dias[di]; return d && d.p === 0 && ind.pmax > 0; }});
    var media = tots.filter(function(v){{return v>0;}}).reduce(function(s,v,_,a){{return s+v/a.length;}}, 0);
    var trend = tot > media ? '\u25b2' : tot < media ? '\u25bc' : '\u2500';
    var trendColor = tot > media ? '#22d3a0' : tot < media ? '#ff4d6d' : '#94a3b8';
    var melhorNome = melhor ? melhor.nome : '-';
    var alertaNomes = alertas.slice(0,2).map(function(a){{return a.nome.split(' ')[0];}}).join(' \u00b7 ');
    kDiv.innerHTML =
      '<div class="ig-kpi" style="border-color:'+tBorder+'">'+
        '<div class="ig-kpi-lbl">TOTAL DE PONTOS</div>'+
        '<div class="ig-kpi-val" style="color:'+tColor+';">'+tot+'<span style="font-size:16px;color:#475569;font-weight:500;"> / '+maxPts+'</span></div>'+
        '<div class="ig-kpi-sub">'+datas[di]+'</div>'+
      '</div>'+
      '<div class="ig-kpi">'+
        '<div class="ig-kpi-lbl">MELHOR INDICADOR</div>'+
        '<div class="ig-kpi-val" style="font-size:16px;color:#e2e8f0;line-height:1.4;letter-spacing:0;padding-top:6px;">'+melhorNome+'</div>'+
        '<div class="ig-kpi-sub" style="color:#22d3a0;font-weight:600;">'+(melhor?melhor.p+' / '+melhor.pmax+' pts obtidos':'')+' </div>'+
      '</div>'+
      '<div class="ig-kpi" style="border-color:rgba(255,77,109,.25);">'+
        '<div class="ig-kpi-lbl">EM ALERTA</div>'+
        '<div class="ig-kpi-val" style="color:#ff4d6d;">'+alertas.length+'<span style="font-size:18px;color:#475569;font-weight:500;"> / 7</span></div>'+
        '<div class="ig-kpi-sub">'+alertaNomes+'</div>'+
      '</div>'+
      '<div class="ig-kpi">'+
        '<div class="ig-kpi-lbl">TEND\u00caNCIA</div>'+
        '<div class="ig-kpi-val" style="color:'+trendColor+';font-size:30px;">'+trend+' '+tot+'</div>'+
        '<div class="ig-kpi-sub">M\u00e9dia: '+media.toFixed(0)+' pts</div>'+
      '</div>';

    // Gauges
    var gDiv = document.getElementById('ig-gauges'); gDiv.innerHTML = '';
    inds.forEach(function(ind) {{
      var d = ind.dias[di] || {{n: null, p: 0}};
      var co = igColorCls(d.p, ind.pmax);
      var arcPct = d.n;
      if (ind.inv && d.n !== null) {{
        var scale = Math.max(ind.m80v * 2, d.n * 1.3);
        arcPct = Math.max(0, Math.min(100, ((scale - d.n) / scale) * 100));
      }}
      var notaReal = d.n !== null ? d.n.toFixed(1)+'%' : '--';
      var card = document.createElement('div');
      card.className = 'ig-gauge-card ' + co.cls;
      card.innerHTML =
        '<div class="ig-gauge-name" style="margin-bottom:4px;">'+ind.nome+'</div>'+
        igGaugeSvg(arcPct, co.c, notaReal, ind.m80v, ind.m100v, ind.inv, d.n) +
        '<div class="ig-gauge-meta" style="margin-top:2px;">Meta: <span style="color:#22d3a0">'+ind.m100+'</span></div>'+
        '<div class="ig-gauge-pts '+co.cls+'">'+d.p+' / '+ind.pmax+' pts</div>';
      gDiv.appendChild(card);
    }});

    // Evolução
    var eDiv = document.getElementById('ig-evo'); eDiv.innerHTML = '';
    var maxTot = Math.max.apply(null, tots.filter(function(v){{return v>0;}})) || 100;
    datas.forEach(function(dt, idx) {{
      var tv = tots[idx] || 0;
      var pct = tv / maxTot;
      var bc = tv >= 85 ? '#22d3a0' : tv >= 60 ? '#fbbf24' : tv > 0 ? '#ff4d6d' : '#1c2237';
      var isAct = idx === di;
      var col = document.createElement('div');
      col.className = 'ig-evo-col';
      col.innerHTML =
        '<div class="ig-evo-num" style="color:'+(isAct?bc:'#475569')+';font-size:'+(isAct?'15':'12')+'px;">'+(tv>0?tv:'')+'</div>'+
        '<div class="ig-evo-bar" style="height:'+Math.max(pct*100,2)+'px;background:'+bc+';opacity:'+(isAct?'1':'0.4')+';'+(isAct?'box-shadow:0 0 14px '+bc+'88;':'')+'"></div>'+
        '<div class="ig-evo-dt" style="color:'+(isAct?'#e8eaf6':'#334155')+';font-weight:'+(isAct?700:400)+';">'+dt+'</div>';
      col.querySelector('.ig-evo-bar').onclick = (function(i){{ return function(){{ igSetDay(col.querySelector('.ig-evo-bar'), i); }}; }})(idx);
      eDiv.appendChild(col);
    }});
  }}
  igRenderSectors();
  </script>
</div>
"""

# --- 3. INJEÇÃO ---
with open(HTML_SRC, "r", encoding="utf-8") as f: html = f.read()

# REMOÇÃO DOS GRÁFICOS DE LINHA DAS OUTRAS ABAS (mantém na Visão Geral)
html = html.replace('<div class="section-title">Evolução da Taxa de Conclusão</div>', '<div class="section-title" style="display:none;"></div>')
html = html.replace('<div class="section-title">Taxa de Conclusão Diária (%)</div>', '<div class="section-title" style="display:none;"></div>')
for canvas_id in ["lineChartTec", "lineChartCit", "lineChart2"]:
    html = html.replace(f'<div class="chart-card" style="margin-bottom:28px;">\n    <canvas id="{canvas_id}"', f'<div class="chart-card" style="display:none;margin-bottom:28px;">\n    <canvas id="{canvas_id}"')
    html = html.replace(f'<div class="chart-card">\n    <canvas id="{canvas_id}"', f'<div class="chart-card" style="display:none;">\n    <canvas id="{canvas_id}"')
    html = html.replace(f'id="{canvas_id}"', f'id="{canvas_id}" style="display:none;"')
    html = html.replace(f'<canvas id="{canvas_id}"', f'<canvas id="{canvas_id}" style="display:none;"')

# WRAPPING do gráfico da Visão Geral com ID controlável
html = html.replace(
    '<div class="section-title">Evolução Diária',
    '<div id="evo-chart-wrapper"><div class="section-title">Evolução Diária'
)
html = html.replace(
    '<canvas id="lineChart" height="80"></canvas>\n  </div>\n</div>',
    '<canvas id="lineChart" height="80"></canvas>\n  </div>\n</div></div>'
)

# MODIFICAR showPage para controlar visibilidade do gráfico
html = html.replace(
    "if (btn) btn.classList.add('active');",
    "if (btn) btn.classList.add('active');\n  var evoW = document.getElementById('evo-chart-wrapper');\n  if(evoW) evoW.style.display = (id==='visao-geral') ? 'block' : 'none';"
)

# ADIÇÃO DE ÍCONES NAS ABAS (MODO PREMIUM)
# Como o template já pode conter os emojis, fazemos o replace usando o texto exato do template
html = html.replace('Evolu\u00e7\u00e3o</button>', 'Evolu\u00e7\u00e3o</button>\n  <button class="nav-btn" onclick="showPage(\'frota\', this)">\U0001f697 Frota</button>\n  <button class="nav-btn" onclick="showPage(\'indicadores\', this)">\U0001f3af Indicadores</button>\n  <button class="nav-btn" onclick="showPage(\'indicador-giga\', this)">\U0001f4ca Indicador Giga</button>\n  <button class="nav-btn" onclick="showPage(\'whatsapp-agenda\', this)">\U0001f4f1 WhatsApp Agenda</button>\n  <button class="nav-btn" onclick="showPage(\'monitor-tv\', this)" style="background:linear-gradient(135deg,#ef4444,#f97316);color:white;">\U0001f4fa Monitor TV</button>\n  <button class="nav-btn" onclick="showPage(\'estoque\', this)">\U0001f4e6 Estoque</button>')

# PÁGINA DO MONITOR TV (iframe)
monitor_tv_page_html = """
<!-- ==================== MONITOR TV ==================== -->
<div class="page" id="monitor-tv">
  <iframe src="monitor_os_tv.html" style="width:100%;height:calc(100vh - 120px);border:none;border-radius:12px;background:#0a0d14;"></iframe>
</div>
"""

# ==================== ESTOQUE (Inventário) ====================
inv_files = glob.glob(os.path.join(PASTA, "inventario_*.xlsx"))
inv_files.sort(key=os.path.getmtime, reverse=True)

estoque_page_html = f"""
<!-- ==================== ESTOQUE ==================== -->
<style>
  .estoque-btn {{ display:inline-flex; align-items:center; gap:8px; background:linear-gradient(135deg, #00e5ff, #7c3aed); color:white; padding:8px 16px; border-radius:8px; text-decoration:none; font-weight:700; font-size:12px; font-family:var(--font-head); transition:transform 0.2s, box-shadow 0.2s; }}
  .estoque-btn:hover {{ transform:translateY(-2px); box-shadow:0 6px 15px rgba(0,229,255,0.3); }}
</style>
<div class="page" id="estoque">
  <div class="kpi-grid" style="grid-template-columns:1fr; margin-bottom:24px;">
    <div class="kpi-card blue" style="display:flex; justify-content:space-between; align-items:center;">
      <div>
        <div class="kpi-label">Base de Inventário</div>
        <div class="kpi-value blue">{len(inv_files)}</div>
        <div class="kpi-sub">relatórios diários salvos no servidor</div>
      </div>
      <i class="fas fa-boxes" style="font-size:40px; color:rgba(0,229,255,0.2);"></i>
    </div>
  </div>
  <div class="section-title"><i class="fas fa-download" style="color:var(--accent)"></i> Download de Inventários</div>
  <div class="chart-card table-scroll" style="max-height:500px;overflow-y:auto;">
    <table class="data-table">
      <thead><tr><th>Data do Relatório</th><th>Tamanho Estimado</th><th>Ação</th></tr></thead>
      <tbody>
"""
for inv in inv_files:
    nome_arq = os.path.basename(inv)
    try: tam_kb = int(os.path.getsize(inv) / 1024)
    except: tam_kb = 0
    data_str = nome_arq.replace("inventario_", "").replace(".xlsx", "").replace("-", "/")
    estoque_page_html += f'<tr><td style="font-weight:700;color:white;font-size:14px;"><i class="far fa-calendar-alt" style="color:var(--muted);margin-right:8px;"></i>{data_str}</td><td><span class="badge" style="background:rgba(255,255,255,0.05)">{tam_kb} KB</span></td><td><a href="inventarios/{nome_arq}" download="{nome_arq}" class="estoque-btn"><i class="fas fa-file-excel"></i> Baixar Planilha</a></td></tr>\n'

estoque_page_html += """
      </tbody>
    </table>
  </div>
</div>
"""

# INJEÇÃO DAS PÁGINAS (antes do footer)
html = html.replace('<div class="footer">', frota_page_html + indicadores_page_html + indicador_giga_page_html + whatsapp_agenda_page_html + monitor_tv_page_html + estoque_page_html + '<div class="footer">') 

# ESTILO DA BARRA DE FILTROS (PREMIUM)
extra_css = """
  .header { border-bottom: none; padding: 20px 32px 10px 32px; }
  .filter-bar { display:flex; align-items:center; justify-content:space-between; padding:4px 32px 20px 32px; background:transparent; gap:12px; position: relative; }
  .filter-group { display:flex; gap:12px; flex:1; }
  .filter-item { 
    display:flex; align-items:center; gap:10px; background:#111520; padding:10px 18px; border-radius:8px; 
    font-size:13px; color:#e8eaf6; border:1px solid #1c2237; cursor: pointer; position: relative; transition: all 0.2s;
    user-select: none;
  }
  .filter-item:hover { border-color: #7c3aed; background: #161c2e; }
  .filter-item i:first-child { color: var(--accent); }
  .filter-item i.fa-chevron-down { font-size: 10px; margin-left: 15px; color: #6b7280; }
  .filter-dropdown { position:absolute; top:100%; left:0; min-width:220px; background:#111520; border:1px solid #1c2237; border-radius:8px; margin-top:8px; box-shadow:0 10px 25px rgba(0,0,0,0.5); display:none; max-height:300px; overflow-y:auto; z-index:1000; }
  .filter-dropdown.show { display:block; }
  .filter-opt { padding:10px 16px; font-size:13px; color:#94a3b8; border-bottom:1px solid rgba(255,255,255,0.03); cursor:pointer; }
  .filter-opt:hover { background:#7c3aed; color:white; }
  .date-filter { display:flex; align-items:center; gap:8px; background:#111520; padding:6px 14px; border-radius:8px; border:1px solid #1c2237; }
  .date-filter label { font-size:11px; color:#6b7280; font-family:var(--font-mono); text-transform:uppercase; }
  .date-filter input[type=date] { background:transparent; border:none; color:#e8eaf6; font-family:var(--font-mono); font-size:13px; outline:none; cursor:pointer; }
  .date-filter input[type=date]::-webkit-calendar-picker-indicator { filter:invert(0.7); cursor:pointer; }
  .date-filter-sep { color:#6b7280; font-size:13px; }
  .date-filter-btn { background:#7c3aed; border:none; color:white; padding:6px 12px; border-radius:6px; font-size:11px; font-family:var(--font-head); font-weight:600; cursor:pointer; transition:all 0.2s; }
  .date-filter-btn:hover { background:#6d28d9; }
  .date-filter-btn.reset { background:transparent; border:1px solid #1c2237; color:#94a3b8; }
  .date-filter-btn.reset:hover { border-color:#7c3aed; color:white; }
  .nav { border-bottom: none !important; border: none !important; background: #0a0d14; padding-bottom: 20px; box-shadow: none !important; }
  .nav::after, .nav::before { display: none !important; }
  .section, .page { border-top: none !important; }
  .nav-btn.active { background: #7c3aed !important; border-radius: 10px; box-shadow: 0 4px 15px rgba(124, 58, 237, 0.3); }
"""
html = html.replace('</style>', extra_css + '</style>')

# REMOÇÃO DA BORDA DA NAV (somente nav, sem afetar outros elementos)
html = html.replace('border-bottom:1px solid var(--border)', 'border-bottom:none')

# DATA DINÂMICA NO CABEÇALHO
m_nomes = ["JAN", "FEV", "MAR", "ABR", "MAI", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
d_nomes = ["SEG", "TER", "QUA", "QUI", "SEX", "SÁB", "DOM"]
agora = datetime.now()
data_header = f'<span class="live-dot"></span> {d_nomes[agora.weekday()]}, {agora.day:02d} DE {m_nomes[agora.month-1]} DE {agora.year}'

# ESCONDER DATA ORIGINAL DO HEADER E INJETAR BARRA DE FILTROS + DATA
html = html.replace('<div class="header-date">', '<div class="header-date" style="display:none;">')

top_struct = f"""
  <div class="filter-bar">
    <div class="filter-group">
      <div class="date-filter">
        <label>De</label>
        <input type="date" id="date-start" onchange="runDateFilter()">
        <span class="date-filter-sep">→</span>
        <label>Até</label>
        <input type="date" id="date-end" onchange="runDateFilter()">
        <button class="date-filter-btn reset" onclick="resetDateFilter()">Limpar</button>
      </div>
      <div class="filter-item" onclick="toggleFiltro('city',this)"><i class="fas fa-globe-americas"></i> <span>Todas Cidades</span> <i class="fas fa-chevron-down"></i><div class="filter-dropdown" id="drop-city"></div></div>
      <div class="filter-item" onclick="toggleFiltro('tec',this)"><i class="fas fa-user-circle"></i> <span>Todos Técnicos</span> <i class="fas fa-chevron-down"></i><div class="filter-dropdown" id="drop-tec"></div></div>
      <div class="filter-item" onclick="toggleFiltro('type',this)"><i class="fas fa-tools"></i> <span>Todos Serviços</span> <i class="fas fa-chevron-down"></i><div class="filter-dropdown" id="drop-type"></div></div>
    </div>
    <div class="header-date">{data_header}</div>
  </div>
"""
html = html.replace('<div class="nav">', top_struct + '<div class="nav">')

# INJEÇÃO DE DADOS NAS VARIÁVEIS JS
html = html.replace("const tecData = [", f"const tecData = {json.dumps(list_tec, ensure_ascii=False)} // [")
html = html.replace("const cityData = [", f"const cityData = {json.dumps(list_city, ensure_ascii=False)} // [")
html = html.replace("const tipoData = [", f"const tipoData = {json.dumps(list_tipo, ensure_ascii=False)} // [")
html = html.replace("const diaData = [", f"const diaData = {json.dumps(list_diag, ensure_ascii=False)} // [")

# INDICADORES DE TEMPO
tempo_medio_os = round(sum(tempos_duracao) / len(tempos_duracao)) if tempos_duracao else 0
tempo_medio_desl = round(sum(tempos_deslocamento) / len(tempos_deslocamento)) if tempos_deslocamento else 0
total_reparos = status_counts["Concluída"]
sla_pct = round(sla_estourado / total_reparos * 100) if total_reparos > 0 else 0

indicadores_tempo_html = f"""
  <div class="section-title"><i class="fas fa-clock" style="color:var(--accent)"></i> Indicadores de Tempo</div>
  <div class="kpi-grid" style="grid-template-columns:repeat(3,1fr);margin-bottom:28px;">
    <div class="kpi-card yellow">
      <div class="kpi-label">Tempo Médio O.S.</div>
      <div class="kpi-value yellow" id="ind-tempo-os">{tempo_medio_os}m</div>
      <div class="kpi-sub">média em minutos</div>
    </div>
    <div class="kpi-card blue">
      <div class="kpi-label">Deslocamento Médio</div>
      <div class="kpi-value blue" id="ind-tempo-desl">{tempo_medio_desl}m</div>
      <div class="kpi-sub">média em minutos</div>
    </div>
    <div class="kpi-card red">
      <div class="kpi-label">SLA Estourado (&gt;24h)</div>
      <div class="kpi-value red" id="ind-sla-val">{sla_estourado}</div>
      <div class="kpi-sub" id="ind-sla-sub">{sla_pct}% dos Servi\u00e7os ({total_reparos})</div>
    </div>
  </div>
"""
# ── SEÇÃO SLA DE ATIVAÇÃO (TV) ──
sla_no_prazo = sorted([s for s in sla_ativacao_list if s["status_sla"] == "no_prazo"], key=lambda x: x["restante_min"])
sla_fora = [s for s in sla_ativacao_list if s["status_sla"] == "fora"]
sla_total_ativ = len(sla_ativacao_list)
sla_total_fora = len(sla_fora)
sla_total_prazo = len(sla_no_prazo)

def _fmt_sla_tempo(minutos):
    h = minutos // 60
    m = minutos % 60
    return f"{h}h {m:02d}m"

sla_cards_html = ""
for s in sla_no_prazo:
    sla_cards_html += f'''
    <div style="background:#0a1628;border:2px solid #22c55e33;border-radius:12px;padding:16px;display:flex;flex-direction:column;gap:8px">
      <div style="display:flex;justify-content:space-between;align-items:flex-start">
        <div style="font-weight:700;color:#e2e8f0;font-size:13px;line-height:1.3;max-width:70%">{s["tec"]}</div>
        <div style="font-size:11px;color:#64748b">OS: {s["os"]}</div>
      </div>
      <div style="font-size:11px;color:#22c55e;font-weight:600">{s["tipo"]}</div>
      <div style="background:#0d1520;border:1px solid #22c55e33;border-radius:8px;padding:12px;text-align:center;margin-top:4px">
        <div style="font-size:11px;color:#94a3b8;margin-bottom:4px">SLA RESTANTE</div>
        <div style="font-size:22px;font-weight:800;color:#22c55e;font-family:var(--font-mono)">{_fmt_sla_tempo(s["restante_min"])}</div>
      </div>
    </div>'''

for s in sla_fora:
    sla_cards_html += f'''
    <div style="background:#0a1628;border:2px solid #ef444433;border-radius:12px;padding:16px;display:flex;flex-direction:column;gap:8px">
      <div style="display:flex;justify-content:space-between;align-items:flex-start">
        <div style="font-weight:700;color:#e2e8f0;font-size:13px;line-height:1.3;max-width:70%">{s["tec"]}</div>
        <div style="font-size:11px;color:#64748b">OS: {s["os"]}</div>
      </div>
      <div style="font-size:11px;color:#ef4444;font-weight:600">{s["tipo"]}</div>
      <div style="background:#1a0a0a;border:1px solid #ef444433;border-radius:8px;padding:12px;text-align:center;margin-top:4px">
        <div style="font-size:11px;color:#94a3b8;margin-bottom:4px">SLA RESTANTE</div>
        <div style="font-size:22px;font-weight:800;color:#ef4444;font-family:var(--font-mono)">FORA DO PRAZO</div>
      </div>
    </div>'''

sla_ativacao_section = f"""
  <div style="margin-top:28px;margin-bottom:28px">
    <div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:16px">
      <div class="section-title" style="margin:0"><i class="fas fa-clock" style="color:#f97316"></i> SLA DE ATIVAÇÃO <span style="background:#f9731622;color:#f97316;padding:4px 12px;border-radius:20px;font-size:14px;font-weight:700;margin-left:8px">{sla_total_ativ}</span></div>
      <div style="display:flex;gap:12px;font-size:13px">
        <span style="color:#22c55e;font-weight:600">✅ {sla_total_prazo} no prazo</span>
        <span style="color:#ef4444;font-weight:600">❌ {sla_total_fora} fora do prazo</span>
      </div>
    </div>
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:12px;max-height:500px;overflow-y:auto;padding-right:4px">
      {sla_cards_html}
    </div>
  </div>
""" if sla_total_ativ > 0 else ""

print(f"[OK] SLA de Ativação: {sla_total_prazo} no prazo, {sla_total_fora} fora do prazo ({sla_total_ativ} total)")

html = html.replace('<div class="grid-2">', indicadores_tempo_html + sla_ativacao_section + '<div class="grid-2">', 1)

# INJEÇÃO DO GRÁFICO DONUT E MOTIVOS
html = html.replace('<div class="donut-wrap" id="donut-container"></div>', generate_donut_svg(status_counts))
html = html.replace('<div class="chart-card" id="motivos-container">', generate_motivos_html(motivos_canc))

# KPIs Visão Geral
html = html.replace('<div class="kpi-value blue">2.851</div>', f'<div class="kpi-value blue">{total_os}</div>')
html = html.replace('<div class="kpi-value green">1.426</div>', f'<div class="kpi-value green">{status_counts["Concluída"]}</div>')
html = html.replace('<div class="kpi-value red">925</div>', f'<div class="kpi-value red">{status_counts["Cancelada"]}</div>')
html = html.replace('<div class="kpi-value yellow">259</div>', f'<div class="kpi-value yellow">{status_counts["Não Concluída"]}</div>')
html = html.replace('<div class="kpi-value white">141</div>', f'<div class="kpi-value white">{status_counts["Suspensa"]}</div>')
html = html.replace('<div class="kpi-value white">100</div>', f'<div class="kpi-value white">{status_counts["Pendente"]}</div>')

# Período real dos dados
datas_unicas = sorted(list(set(d["date"] for d in raw_os_data if d["date"] != "--")))
periodo_sub = f"{datas_unicas[0]} a {datas_unicas[-1]}" if datas_unicas else "sem dados"
html = html.replace('últimos 8 dias úteis', periodo_sub)
html = html.replace('Últimos 8 dias úteis', periodo_sub)

# Recordes IDs
reps = {"top-tec-taxa": f"{v_tec_taxa['taxa']}%", "top-tec-nome": v_tec_taxa['Recurso'], "top-tec-prod": v_tec_prod['total'], "top-tec-prod-nome": v_tec_prod['Recurso'], "bad-tec-taxa": f"{v_tec_bad['taxa']}%", "bad-tec-nome": v_tec_bad['Recurso'], "top-cit-taxa": f"{v_cit_taxa['taxa']}%", "top-cit-nome": v_cit_taxa['Cidade'], "top-cit-vol": v_cit_vol['total'], "top-cit-vol-nome": v_cit_vol['Cidade'], "bad-cit-taxa": f"{v_cit_bad['taxa']}%", "bad-cit-nome": v_cit_bad['Cidade'], "evo-pico-vol": v_pico['total'], "evo-pico-data": v_pico['data'], "evo-top-taxa": f"{v_melhor['taxa']}%", "evo-top-data": v_melhor['data'], "evo-bad-taxa": f"{v_pior['taxa']}%", "evo-bad-data": v_pior['data'], "evo-avg-vol": f"{v_media:.1f}"}
for k, v in reps.items(): html = html.replace(f'id="{k}">0<', f'id="{k}">{v}<').replace(f'id="{k}">0%<', f'id="{k}">{v}<').replace(f'id="{k}">-<', f'id="{k}">{v}<')

# INJEÇÃO DE DADOS BRUTOS E LISTAS PARA FILTROS
filtros_listas = {
    "periodos": ["Hoje", "\u00daltimos 3 dias", "\u00daltimos 5 dias", "\u00daltimos 7 dias", "\u00daltimos 10 dias"],
    "allDates": datas_unicas,
    "citys": sorted(list(set(d["city"] for d in raw_os_data))),
    "tecs": sorted(list(set(d["tec"] for d in raw_os_data))),
    "types": sorted(list(set(d["type"] for d in raw_os_data)))
}
js_data = "\nvar rawOSData = " + json.dumps(raw_os_data, ensure_ascii=False) + ";\n"
js_data += "var filterLists = " + json.dumps(filtros_listas, ensure_ascii=False) + ";\n"

js_logic = r"""
var filtrosAtivos = {periodo:'Todos', city:'Todos', tec:'Todos', type:'Todos'};
function toggleFiltro(tipo, el) {
  var drop = el.querySelector('.filter-dropdown');
  var wasOpen = drop.classList.contains('show');
  document.querySelectorAll('.filter-dropdown').forEach(function(d){d.classList.remove('show');});
  if(wasOpen) return;
  if(drop.innerHTML === '') {
    var defaultLabel = tipo==='city'?'Todas Cidades':tipo==='tec'?'Todos T\u00e9cnicos':tipo==='type'?'Todos Servi\u00e7os':'Todo o Per\u00edodo';
    var optHtml = '<div class="filter-opt" onclick="event.stopPropagation();selFiltro(\''+tipo+'\',\'Todos\')">' + defaultLabel + '</div>';
    var lista = filterLists[tipo + 's'] || [];
    for(var i=0;i<lista.length;i++) {
      var safe = lista[i].replace(/'/g, "\\'");
      optHtml += '<div class="filter-opt" onclick="event.stopPropagation();selFiltro(\''+tipo+'\',\''+safe+'\')">' + lista[i] + '</div>';
    }
    drop.innerHTML = optHtml;
  }
  drop.classList.add('show');
}
function selFiltro(tipo, val) {
  filtrosAtivos[tipo] = val;
  var item = document.querySelector('[onclick*="toggleFiltro(\''+tipo+'\'"]');
  var defaultLabel = tipo==='city'?'Todas Cidades':tipo==='tec'?'Todos T\u00e9cnicos':tipo==='type'?'Todos Servi\u00e7os':'Todo o Per\u00edodo';
  item.querySelector('span').textContent = val === 'Todos' ? defaultLabel : (val.length>20?val.substring(0,18)+'...':val);
  item.style.borderColor = val === 'Todos' ? '#1c2237' : '#7c3aed';
  document.querySelectorAll('.filter-dropdown').forEach(function(d){d.classList.remove('show');});
  runFilter();
}
function runDateFilter() { runFilter(); }
function resetDateFilter() {
  document.getElementById('date-start').value = '';
  document.getElementById('date-end').value = '';
  runFilter();
}
function runFilter() {
  var ds = document.getElementById('date-start').value;
  var de = document.getElementById('date-end').value;
  var data = rawOSData.filter(function(d) {
    var dateOk = true;
    if(ds || de) {
      var parts = d.date.split('/');
      if(parts.length===2) {
        var yr = new Date().getFullYear();
        var dISO = yr + '-' + parts[1] + '-' + parts[0];
        if(ds && dISO < ds) dateOk = false;
        if(de && dISO > de) dateOk = false;
      }
    }
    return dateOk &&
           (filtrosAtivos.city==='Todos' || d.city===filtrosAtivos.city) &&
           (filtrosAtivos.tec==='Todos' || d.tec===filtrosAtivos.tec) &&
           (filtrosAtivos.type==='Todos' || d.type===filtrosAtivos.type);
  });
  var counts = {'Conclu\u00edda':0,'Cancelada':0,'N\u00e3o Conclu\u00edda':0,'Suspensa':0,'Pendente':0};
  var tecs={}, cities={}, types={}, motivos={};
  data.forEach(function(d) {
    counts[d.status]++;
    if(d.motivo && (d.status==='Cancelada'||d.status==='N\u00e3o Conclu\u00edda')) motivos[d.motivo]=(motivos[d.motivo]||0)+1;
    [{o:tecs,k:d.tec},{o:cities,k:d.city},{o:types,k:d.type}].forEach(function(x){
      if(!x.o[x.k]) x.o[x.k]={key:x.k,total:0,concluido:0,nao_conc:0,cancelado:0};
      x.o[x.k].total++;
      if(d.status==='Conclu\u00edda') x.o[x.k].concluido++;
      else if(d.status==='N\u00e3o Conclu\u00edda') x.o[x.k].nao_conc++;
      else if(d.status==='Cancelada') x.o[x.k].cancelado++;
    });
  });
  // Indicadores de Tempo filtrados
  var totalDur=0, cntDur=0, totalDesl=0, cntDesl=0, slaCount=0, concCount=counts['Conclu\u00edda'];
  data.forEach(function(d){
    if(d.dur>0){totalDur+=d.dur;cntDur++;}
    if(d.desl>0){totalDesl+=d.desl;cntDesl++;}
    if(d.sla===1) slaCount++;
  });
  var avgDur=cntDur>0?Math.round(totalDur/cntDur):0;
  var avgDesl=cntDesl>0?Math.round(totalDesl/cntDesl):0;
  var slaPct=concCount>0?Math.round(slaCount/concCount*100):0;
  var eDur=document.getElementById('ind-tempo-os');if(eDur)eDur.textContent=avgDur+'m';
  var eDesl=document.getElementById('ind-tempo-desl');if(eDesl)eDesl.textContent=avgDesl+'m';
  var eSla=document.getElementById('ind-sla-val');if(eSla)eSla.textContent=slaCount;
  var slaLabel = filtrosAtivos.type!=='Todos' ? filtrosAtivos.type : 'Servi\u00e7os';
  var eSlaSub=document.getElementById('ind-sla-sub');if(eSlaSub)eSlaSub.textContent=slaPct+'% dos '+slaLabel+' ('+concCount+')';
  // KPIs - por classe de cor (mais confiável)
  var vBlue = document.querySelector('.kpi-value.blue');
  var vGreen = document.querySelector('.kpi-value.green');
  var vRed = document.querySelector('.kpi-value.red');
  var vYellow = document.querySelector('.kpi-value.yellow');
  var vWhites = document.querySelectorAll('.kpi-value.white');
  if(vBlue) vBlue.textContent = data.length;
  if(vGreen) vGreen.textContent = counts['Conclu\u00edda'];
  if(vRed) vRed.textContent = counts['Cancelada'];
  if(vYellow) vYellow.textContent = counts['N\u00e3o Conclu\u00edda'];
  if(vWhites.length>=1) vWhites[0].textContent = counts['Suspensa'];
  if(vWhites.length>=2) vWhites[1].textContent = counts['Pendente'];
  // Taxa no card de Concluídas
  var greenCard = document.querySelector('.kpi-value.green');
  if(greenCard) {
    var taxaGeral = data.length>0 ? (counts['Conclu\u00edda']/data.length*100).toFixed(1) : 0;
    var sub = greenCard.parentElement.querySelector('.kpi-sub');
    if(sub) sub.textContent = 'taxa: ' + taxaGeral + '%';
  }
  // Recordes de Técnicos
  var tecsSorted = Object.values(tecs).map(function(x){x.taxa=x.total>0?Math.round(x.concluido/x.total*1000)/10:0;return x;});
  var bestTec = tecsSorted.sort(function(a,b){return b.taxa-a.taxa;})[0];
  var prodTec = tecsSorted.sort(function(a,b){return b.total-a.total;})[0];
  var worstTec = tecsSorted.sort(function(a,b){return a.taxa-b.taxa;})[0];
  if(bestTec){var e=document.getElementById('top-tec-taxa');if(e)e.textContent=bestTec.taxa+'%';var n=document.getElementById('top-tec-nome');if(n)n.textContent=bestTec.key;}
  if(prodTec){var e=document.getElementById('top-tec-prod');if(e)e.textContent=prodTec.total;var n=document.getElementById('top-tec-prod-nome');if(n)n.textContent=prodTec.key;}
  if(worstTec){var e=document.getElementById('bad-tec-taxa');if(e)e.textContent=worstTec.taxa+'%';var n=document.getElementById('bad-tec-nome');if(n)n.textContent=worstTec.key;}
  // Donut
  var dc = document.querySelector('.donut-wrap');
  if(dc) {
    var total = Object.values(counts).reduce(function(a,b){return a+b;},0);
    var colors = ['#22d3a0','#ff4d6d','#fbbf24','#7c3aed','#6b7280'];
    var labels = ['Conclu\u00edda','Cancelada','N\u00e3o Conclu\u00edda','Suspensa','Pendente'];
    var svg = '<svg class="donut-svg" width="120" height="120" viewBox="0 0 42 42"><circle cx="21" cy="21" r="15.915" fill="transparent" stroke="rgba(255,255,255,0.05)" stroke-width="3"></circle>';
    var offset = 25;
    for(var i=0;i<labels.length;i++){
      var p = total>0?(counts[labels[i]]/total*100):0;
      if(p>0){svg+='<circle cx="21" cy="21" r="15.915" fill="transparent" stroke="'+colors[i]+'" stroke-width="3" stroke-dasharray="'+p+' '+(100-p)+'" stroke-dashoffset="'+offset+'"></circle>';offset-=p;}
    }
    svg += '</svg>';
    var leg = '<div class="donut-legend">';
    for(var i=0;i<labels.length;i++){
      var p = total>0?(counts[labels[i]]/total*100).toFixed(1):0;
      leg+='<div class="legend-item"><div class="legend-dot" style="background:'+colors[i]+'"></div>'+labels[i]+'<div class="legend-val">'+counts[labels[i]]+' ('+p+'%)</div></div>';
    }
    dc.innerHTML = svg + leg + '</div>';
  }
  // Motivos
  var mc = document.getElementById('motivos-container');
  if(mc) {
    var sorted = Object.entries(motivos).sort(function(a,b){return b[1]-a[1];}).slice(0,6);
    var mhtml = '<div class="chart-title">Top Motivos de N\u00e3o Conclus\u00e3o / Cancelamento</div>';
    if(sorted.length===0) mhtml+='<div style="color:var(--muted);font-size:12px;height:150px;display:flex;align-items:center;justify-content:center;">Nenhum motivo</div>';
    else { var mx=sorted[0][1]; for(var i=0;i<sorted.length;i++){var w=(sorted[i][1]/mx)*100;mhtml+='<div class="bar-row"><div class="bar-label wide">'+sorted[i][0]+'</div><div class="bar-track"><div class="bar-fill red" style="width:'+w+'%"></div></div><div class="bar-val">'+sorted[i][1]+'</div></div>';}}
    mc.innerHTML = mhtml;
  }
  // Ordena por taxa decrescente para ranking
  var tecList = Object.values(tecs).map(function(x){x.taxa=x.total>0?Math.round(x.concluido/x.total*1000)/10:0;return x;}).sort(function(a,b){return b.taxa-a.taxa;});
  var cityList = Object.values(cities).map(function(x){x.taxa=x.total>0?Math.round(x.concluido/x.total*1000)/10:0;return x;}).sort(function(a,b){return b.taxa-a.taxa;});
  var typeList = Object.values(types).map(function(x){x.taxa=x.total>0?Math.round(x.concluido/x.total*1000)/10:0;return x;}).sort(function(a,b){return b.total-a.total;});
  updateTbl('tec-table-body', tecList);
  updateTbl('city-table-body', cityList);
  updateTbl('tipo-table-body', typeList);
}
function updateTbl(id, list) {
  var tbody = document.getElementById(id);
  if(!tbody) return;
  var rows = '';
  var trofeus = ['\ud83e\udd47','\ud83e\udd48','\ud83e\udd49'];
  for(var i=0;i<Math.min(list.length,20);i++) {
    var item = list[i];
    var taxa = item.taxa || (item.total>0 ? Math.round(item.concluido/item.total*1000)/10 : 0);
    var cls = taxa>80?'green':taxa>50?'yellow':'red';
    var rank = i<3 ? trofeus[i] : (i+1);
    var taxaColor = taxa>=90?'#22d3a0':taxa>=70?'#22d3a0':taxa>=50?'#fbbf24':'#ff4d6d';
    rows += '<tr><td>'+rank+'</td><td>'+item.key+'</td><td>'+item.total+'</td><td style="color:#22d3a0">'+item.concluido+'</td><td style="color:#fbbf24">'+(item.nao_conc||0)+'</td><td style="color:#ff4d6d">'+(item.cancelado||0)+'</td><td><span style="background:#0d1520;border:1px solid '+taxaColor+'33;padding:4px 12px;border-radius:20px;color:'+taxaColor+';font-size:12px;font-family:var(--font-mono)">'+taxa.toFixed(1)+'%</span></td><td><div style="display:flex;align-items:center;gap:10px"><div class="progress-bar" style="flex:1"><div class="progress-fill '+cls+'" style="width:'+taxa+'%"></div></div><span style="color:#94a3b8;font-size:12px;font-family:var(--font-mono);white-space:nowrap">'+taxa.toFixed(1)+'%</span></div></td></tr>';
  }
  tbody.innerHTML = rows;
}
document.addEventListener('click', function(e) {
  if(!e.target.closest('.filter-item')) document.querySelectorAll('.filter-dropdown').forEach(function(d){d.classList.remove('show');});
});
"""
html = html.replace('</script>', js_data + js_logic + '</script>')

with open(HTML_OUT, "w", encoding="utf-8") as f: f.write(html)

def fazer_upload():
    config_file = os.path.join(PASTA, 'config_servidor.json')
    if not os.path.exists(config_file): return
    with open(config_file, 'r', encoding='utf-8') as f: cfg = json.load(f)
    try:
        import paramiko
        client = paramiko.SSHClient(); client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        
        # Usa chave SSH para autenticacao (igual a TV)
        ssh_key_path = os.path.join(os.environ["USERPROFILE"], ".ssh", "id_rsa")
        pkey = paramiko.RSAKey.from_private_key_file(ssh_key_path)
        
        client.connect(cfg['HOST'], port=int(cfg['PORT']), username=cfg['USER'], pkey=pkey)
        sftp = client.open_sftp(); sftp.put(HTML_OUT, os.path.join(cfg['REMOTE_DIR'], "index.html").replace("\\","/"))
        
        # --- UPLOAD DOS INVENTÁRIOS ---
        remote_inv_dir = os.path.join(cfg['REMOTE_DIR'], "inventarios").replace("\\","/")
        try:
            sftp.stat(remote_inv_dir)
        except IOError:
            sftp.mkdir(remote_inv_dir)
            
        inv_files = glob.glob(os.path.join(PASTA, "inventario_*.xlsx"))
        inv_files.sort(key=os.path.getmtime, reverse=True)
        for inv in inv_files[:20]:  # Upando os ultimos 20 dias no maximo
            nome_arq = os.path.basename(inv)
            remote_path = remote_inv_dir + "/" + nome_arq
            try:
                sftp.stat(remote_path)
            except IOError:
                print(f"[SFTP] Enviando novo inventario: {nome_arq}")
                sftp.put(inv, remote_path)
                
        sftp.close(); client.close(); print("[SUCESSO] Dashboard Restaurado 100%!")
        
        # --- PIPELINE DE INJECAO POS-UPLOAD ---
        print("[...] Rodando pipeline de injecao no servidor...")
        try:
            stdin, stdout, stderr = client2 if False else (None, None, None)
        except: pass
        try:
            client2 = paramiko.SSHClient(); client2.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client2.connect(cfg['HOST'], port=int(cfg['PORT']), username=cfg['USER'], pkey=pkey)
            cmd = "cd /docker/dashboard && python3 injetar_auth.py 2>&1 && python3 injetar_admin_btn.py 2>&1 && python3 injetar_templates.py 2>&1 && python3 inject_estoque.py 2>&1 && python3 injetar_melhorias.py 2>&1"
            stdin, stdout, stderr = client2.exec_command(cmd, timeout=120)
            out = stdout.read().decode('utf-8', errors='replace')
            print(out)
            client2.close()
            print("[SUCESSO] Pipeline de injecao concluido!")
        except Exception as e2:
            print(f"[AVISO] Pipeline de injecao falhou: {e2}")
            print("         Os injects serao aplicados automaticamente as 22h pelo servidor.")
        
    except Exception as e: print(f"[ERRO] {e}")

fazer_upload()
