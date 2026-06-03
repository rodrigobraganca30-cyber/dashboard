#!/usr/bin/env python3
"""
injetar_agenda_futura.py
========================
Injeção CIRÚRGICA da aba Agenda Futura no index.html.
NÃO regenera o dashboard inteiro — só substitui a seção entre
<!-- BEGIN AGENDA FUTURA --> e <!-- END AGENDA FUTURA -->.
Se os marcadores não existem, insere antes de <div class="footer">.
"""
import os, sys, json, shutil, datetime
from collections import Counter

INDEX = '/docker/dashboard/html/index.html'
AGENDA_FILE = '/docker/dashboard/html/AGENDA_FUTURA.xlsx'
BEGIN_MARKER = '<!-- BEGIN AGENDA FUTURA -->'
END_MARKER = '<!-- END AGENDA FUTURA -->'


def log(msg):
    ts = datetime.datetime.now().strftime('%H:%M:%S')
    print(f"[{ts}] {msg}")


def ler_agenda_futura(path):
    """Lê AGENDA_FUTURA.xlsx e retorna (dias_dict, tipos_dict, cidades_dict, total, ignorados)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Detecta colunas pelo cabeçalho
    headers = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1)):
        if cell.value:
            headers[str(cell.value).strip().lower()] = cell.column - 1

    col_data = headers.get('data_origem', headers.get('data', None))
    # "Tipo de Atividade_2" é o tipo real (Ativação, Reparo, etc.)
    # "Tipo de Atividade" é o tipo genérico (Normal, Urgente)
    col_tipo = headers.get('tipo de atividade_2',
               headers.get('tipo_atividade_2',
               headers.get('tipo de atividade',
               headers.get('tipo_atividade',
               headers.get('tipo', None)))))
    col_cidade = headers.get('cidade', None)
    col_status = headers.get('status da atividade',
                 headers.get('status_atividade',
                 headers.get('status', None)))

    if col_data is None:
        raise ValueError(f"Coluna 'data_origem'/'data' não encontrada. Headers: {list(headers.keys())}")

    dias = {}       # {"31/05": {"total": N, "tipos": Counter, "cidades": Counter, "data_iso": "2026-05-31"}}
    tipos = {}      # {"Ativação": N}
    cidades = {}    # {"MACAE": N}
    total = 0
    ignorados = 0
    hoje = datetime.date.today()

    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = list(row)

        # Data
        data_val = vals[col_data] if col_data < len(vals) else None
        if not data_val:
            continue

        if isinstance(data_val, datetime.datetime):
            data_obj = data_val.date()
        elif isinstance(data_val, datetime.date):
            data_obj = data_val
        elif isinstance(data_val, str):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    data_obj = datetime.datetime.strptime(data_val.strip(), fmt).date()
                    break
                except ValueError:
                    continue
            else:
                continue
        else:
            continue

        # Só futuras
        if data_obj <= hoje:
            ignorados += 1
            continue

        # Status — ignora canceladas/admin
        status = str(vals[col_status]).strip().upper() if col_status is not None and col_status < len(vals) and vals[col_status] else ""
        if any(x in status for x in ['CANCEL', 'ADMIN', 'DELETE']):
            ignorados += 1
            continue

        tipo = str(vals[col_tipo]).strip() if col_tipo is not None and col_tipo < len(vals) and vals[col_tipo] else "Outros"
        cidade = str(vals[col_cidade]).strip() if col_cidade is not None and col_cidade < len(vals) and vals[col_cidade] else "N/I"

        # Ignora tipos internos que não são OS de campo
        tipo_upper = tipo.upper()
        if any(x in tipo_upper for x in ['ALMOÇO', 'ALMOCO', 'CHECKLIST', 'CHECK LIST', 'LUNCH']):
            ignorados += 1
            continue

        dia_key = data_obj.strftime('%d/%m')
        data_iso = data_obj.isoformat()

        if dia_key not in dias:
            dias[dia_key] = {"total": 0, "tipos": Counter(), "cidades": Counter(), "data_iso": data_iso}
        dias[dia_key]["total"] += 1
        dias[dia_key]["tipos"][tipo] += 1
        dias[dia_key]["cidades"][cidade] += 1

        tipos[tipo] = tipos.get(tipo, 0) + 1
        cidades[cidade] = cidades.get(cidade, 0) + 1
        total += 1

    wb.close()
    return dias, tipos, cidades, total, ignorados


def gerar_html_secao(dias, tipos, cidades, total):
    """Gera o HTML completo da seção Agenda Futura (idêntico ao gerar_dashboard_v2.py)."""

    dias_count = len(dias)
    media_dia = round(total / dias_count, 1) if dias_count > 0 else 0
    pico_label = max(dias, key=lambda k: dias[k]["total"]) if dias else "-"
    pico_val = dias[pico_label]["total"] if pico_label != "-" else 0
    top_tipo = max(tipos, key=tipos.get) if tipos else "-"
    top_tipo_qtd = tipos.get(top_tipo, 0)

    # Dados para gráfico
    dias_sorted = sorted(dias.items(), key=lambda x: x[1].get("data_iso", x[0]))
    chart_labels = json.dumps([d[0] for d in dias_sorted], ensure_ascii=False)
    chart_values = json.dumps([d[1]["total"] for d in dias_sorted])

    # Barras de tipos (top 8)
    tipos_sorted = sorted(tipos.items(), key=lambda x: x[1], reverse=True)[:8]
    tipo_max = tipos_sorted[0][1] if tipos_sorted else 1
    tipo_bars = ""
    for nome, qtd in tipos_sorted:
        pct = round(qtd / tipo_max * 100)
        tipo_bars += f'<div class="bar-row"><div class="bar-label wide">{nome}</div><div class="bar-track"><div class="bar-fill blue" style="width:{pct}%"></div></div><div class="bar-val">{qtd}</div></div>'

    # Barras de cidades (top 10)
    cidades_sorted = sorted(cidades.items(), key=lambda x: x[1], reverse=True)[:10]
    cid_max = cidades_sorted[0][1] if cidades_sorted else 1
    cid_bars = ""
    for nome, qtd in cidades_sorted:
        pct = round(qtd / cid_max * 100)
        cid_bars += f'<div class="bar-row"><div class="bar-label wide">{nome}</div><div class="bar-track"><div class="bar-fill" style="width:{pct}%;background:linear-gradient(90deg,#7c3aed,#00e5ff)"></div></div><div class="bar-val">{qtd}</div></div>'

    # Tabela diária
    table_rows = ""
    for d_key, d_val in dias_sorted:
        top3 = ", ".join([f"{t}: {q}" for t, q in d_val["tipos"].most_common(3)])
        top1_cidade = d_val["cidades"].most_common(1)[0][0] if d_val["cidades"] else "-"
        n_cidades = len(d_val["cidades"])
        extra = f' <span style="color:var(--muted);font-weight:400">+{n_cidades-1} cidades</span>' if n_cidades > 1 else ''
        table_rows += (
            f'<tr>'
            f'<td style="font-family:var(--font-mono);font-weight:600;color:var(--accent)">{d_key}</td>'
            f'<td style="font-family:var(--font-mono);font-size:18px;font-weight:800;color:var(--text)">{d_val["total"]}</td>'
            f'<td style="font-size:12px;color:var(--muted)">{top3}</td>'
            f'<td style="font-size:12px;font-weight:600">{top1_cidade}{extra}</td>'
            f'</tr>'
        )

    # Dados JSON para exportação
    export_data = []
    for d_key, d_val in dias_sorted:
        export_data.append({
            "data": d_key,
            "total": d_val["total"],
            "data_iso": d_val.get("data_iso", ""),
            "tipos": dict(d_val["tipos"].most_common()),
            "cidades": dict(d_val["cidades"].most_common()),
        })
    export_json = json.dumps(export_data, ensure_ascii=False)

    sem_dados = "" if total > 0 else '<div style="text-align:center;padding:80px 20px;color:var(--muted);font-size:14px;">⏳ Dados da Agenda Futura ainda não disponíveis.<br>O robô roda todos os dias às 05:00 AM e irá preencher este painel automaticamente.</div>'

    tipo_content = tipo_bars if total > 0 else '<div style="color:var(--muted);font-size:13px;text-align:center;padding:40px">Sem dados</div>'
    cid_content = cid_bars if total > 0 else '<div style="color:var(--muted);font-size:13px;text-align:center;padding:40px">Sem dados</div>'
    table_content = table_rows if total > 0 else '<tr><td colspan="4" style="text-align:center;color:var(--muted);padding:40px">Dados não disponíveis</td></tr>'
    kpi_display = "" if total > 0 else 'style="display:none"'

    agora = datetime.datetime.now().strftime('%d/%m/%Y %H:%M')

    html = f"""{BEGIN_MARKER}
<!-- ==================== AGENDA FUTURA ==================== -->
<style>
  .af-export-btn {{
    display:inline-flex; align-items:center; gap:8px;
    background:linear-gradient(135deg,#7c3aed,#00e5ff);
    border:none; color:white; padding:10px 22px; border-radius:8px;
    font-family:var(--font-head); font-size:13px; font-weight:700;
    cursor:pointer; transition:all .2s; text-decoration:none;
    box-shadow:0 4px 15px rgba(124,58,237,0.4);
  }}
  .af-export-btn:hover {{ opacity:.85; transform:translateY(-1px); }}
  .af-header-row {{
    display:flex; align-items:center; justify-content:space-between; margin-bottom:24px;
  }}
  .af-update-info {{
    font-family:var(--font-mono); font-size:11px; color:var(--muted);
  }}
</style>
<div class="page" id="agenda-futura">
  <div class="af-header-row">
    <div>
      <div class="section-title" style="margin-bottom:4px;">🔮 Agenda Futura — Próximos 30 dias</div>
      <div class="af-update-info">Atualizado pelo robô às {agora} · Dados sem atribuição de técnico (por cidade)</div>
    </div>
    <button class="af-export-btn" onclick="exportarAgendaFuturaCSV()" id="btn-exportar-agenda">
      ⬇ Exportar Planilha
    </button>
  </div>

  {sem_dados}

  <div class="kpi-grid" style="grid-template-columns:repeat(4,1fr);margin-bottom:28px;" {kpi_display}>
    <div class="kpi-card blue">
      <div class="kpi-label">Total de OS Futuras</div>
      <div class="kpi-value blue">{total:,}</div>
      <div class="kpi-sub">{dias_count} dias mapeados</div>
    </div>
    <div class="kpi-card green">
      <div class="kpi-label">Média por Dia</div>
      <div class="kpi-value green">{media_dia}</div>
      <div class="kpi-sub">OS/dia agendadas</div>
    </div>
    <div class="kpi-card yellow">
      <div class="kpi-label">Pico de Agendamento</div>
      <div class="kpi-value yellow">{pico_val}</div>
      <div class="kpi-sub">dia {pico_label}</div>
    </div>
    <div class="kpi-card purple">
      <div class="kpi-label">Tipo com Maior Volume</div>
      <div class="kpi-value white" style="font-size:18px;line-height:1.2">{top_tipo}</div>
      <div class="kpi-sub">{top_tipo_qtd} O.S.</div>
    </div>
  </div>

  <div class="section-title">📅 Carga Diária — Próximos 30 Dias</div>
  <div class="chart-card" style="margin-bottom:24px;">
    <canvas id="afTimelineChart" height="220"></canvas>
  </div>

  <div class="grid-2" style="margin-bottom:24px;">
    <div class="chart-card">
      <div class="chart-title">🔧 Volume por Tipo de Serviço</div>
      <div id="af-tipo-bars">{tipo_content}</div>
    </div>
    <div class="chart-card">
      <div class="chart-title">🗺️ Volume por Cidade (Top 10)</div>
      <div id="af-cidade-bars">{cid_content}</div>
    </div>
  </div>

  <div class="section-title">📋 Detalhamento por Dia</div>
  <div class="chart-card table-scroll">
    <table class="data-table" id="af-table-dias">
      <thead>
        <tr>
          <th>Data</th>
          <th>Total OS</th>
          <th>Top Serviços</th>
          <th>Principal Cidade</th>
        </tr>
      </thead>
      <tbody>
        {table_content}
      </tbody>
    </table>
  </div>

  <script>
  var agendaFuturaData = {export_json};
  var _afChartCreated = false;
  var _afLabels = {chart_labels};
  var _afValues = {chart_values};
  function _afInitChart() {{
    if (_afChartCreated) return;
    if (typeof Chart === 'undefined') {{ setTimeout(_afInitChart, 300); return; }}
    var ctx = document.getElementById('afTimelineChart');
    if (!ctx) return;
    if (ctx.offsetWidth === 0) return;
    _afChartCreated = true;
    new Chart(ctx, {{
      type: 'line',
      data: {{
        labels: _afLabels,
        datasets: [{{
          label: 'OS Agendadas',
          data: _afValues,
          borderColor: '#7c3aed',
          backgroundColor: 'rgba(124,58,237,0.12)',
          borderWidth: 2.5,
          tension: 0.4,
          fill: true,
          pointBackgroundColor: '#7c3aed',
          pointRadius: 4,
          pointHoverRadius: 7
        }}]
      }},
      options: {{
        responsive: true,
        plugins: {{
          legend: {{ display: false }},
          tooltip: {{
            backgroundColor: '#161c2e',
            borderColor: 'rgba(255,255,255,.1)',
            borderWidth: 1,
            titleFont: {{family:'DM Mono',size:12}},
            bodyFont: {{family:'DM Mono',size:12}},
            titleColor: '#e8eaf6',
            bodyColor: '#94a3b8',
            padding: 10
          }}
        }},
        scales: {{
          x: {{ grid: {{color:'rgba(255,255,255,.05)'}}, ticks: {{color:'#6b7280',font:{{family:'DM Mono',size:10}},maxRotation:45}} }},
          y: {{ grid: {{color:'rgba(255,255,255,.05)'}}, ticks: {{color:'#6b7280',font:{{family:'DM Mono',size:11}}}} }}
        }}
      }}
    }});
  }}
  document.addEventListener('DOMContentLoaded', function() {{
    var _origShowPage = window.showPage;
    if (_origShowPage) {{
      window.showPage = function(id, btn) {{
        _origShowPage(id, btn);
        if (id === 'agenda-futura') {{ setTimeout(_afInitChart, 200); }}
      }};
    }}
    var afPage = document.getElementById('agenda-futura');
    if (afPage && afPage.classList.contains('active')) {{ setTimeout(_afInitChart, 500); }}
  }});
  function exportarAgendaFuturaCSV() {{
    var btn = document.getElementById('btn-exportar-agenda');
    if (btn) {{ btn.textContent = '⏳ Baixando...'; btn.disabled = true; }}
    var a = document.createElement('a');
    a.href = 'AGENDA_FUTURA.xlsx';
    a.download = 'AGENDA_FUTURA.xlsx';
    document.body.appendChild(a);
    a.click();
    document.body.removeChild(a);
    setTimeout(function() {{
      if (btn) {{ btn.textContent = '⬇ Exportar Planilha'; btn.disabled = false; }}
    }}, 2000);
  }}
  </script>
</div>
{END_MARKER}"""

    return html


def main():
    # 1. Lê AGENDA_FUTURA.xlsx
    if not os.path.exists(AGENDA_FILE):
        log(f"ERRO: {AGENDA_FILE} não encontrado!")
        sys.exit(1)

    log(f"Lendo {AGENDA_FILE}...")
    try:
        dias, tipos, cidades, total, ignorados = ler_agenda_futura(AGENDA_FILE)
        log(f"OK: {total} OS | {len(dias)} dias | {len(cidades)} cidades | {ignorados} ignoradas")
    except Exception as e:
        log(f"ERRO ao ler AGENDA_FUTURA.xlsx: {e}")
        sys.exit(1)

    # 2. Gera HTML da seção
    secao_html = gerar_html_secao(dias, tipos, cidades, total)
    log(f"HTML da seção gerado: {len(secao_html)} bytes")

    # 3. Lê index.html
    if not os.path.exists(INDEX):
        log(f"ERRO: {INDEX} não encontrado!")
        sys.exit(1)

    with open(INDEX, 'r', encoding='utf-8') as f:
        html = f.read()

    tamanho_antes = len(html)
    log(f"index.html lido: {tamanho_antes:,} bytes")

    # 4. Backup
    ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = f"{INDEX}.bak.agenda_{ts}"
    shutil.copy2(INDEX, bak)
    log(f"Backup: {bak}")

    # 5. Substitui ou insere a seção
    if BEGIN_MARKER in html and END_MARKER in html:
        # Replace entre marcadores
        start = html.index(BEGIN_MARKER)
        end = html.index(END_MARKER) + len(END_MARKER)
        html = html[:start] + secao_html + html[end:]
        log("Seção Agenda Futura SUBSTITUÍDA (entre marcadores)")
    elif '<div class="footer">' in html:
        # Insere antes do footer
        html = html.replace('<div class="footer">', secao_html + '\n<div class="footer">')
        log("Seção Agenda Futura INSERIDA (antes do footer)")
    else:
        # Insere antes de </body>
        html = html.replace('</body>', secao_html + '\n</body>')
        log("Seção Agenda Futura INSERIDA (antes de </body>)")

    # 6. Verifica integridade — confere que as outras abas continuam
    abas_criticas = ['visao-geral', 'tecnicos', 'cidades', 'servicos', 'evolucao',
                     'frota', 'indicadores', 'indicador-giga', 'whatsapp-agenda',
                     'monitor-tv', 'estoque']
    abas_ok = True
    for aba in abas_criticas:
        if f'id="{aba}"' not in html:
            log(f"⚠️ ALERTA: aba '{aba}' DESAPARECEU! Abortando!")
            abas_ok = False

    if not abas_ok:
        log("ERRO FATAL: Abas críticas desapareceram. Restaurando backup...")
        shutil.copy2(bak, INDEX)
        sys.exit(1)

    if 'id="agenda-futura"' not in html:
        log("ERRO: Seção agenda-futura não foi inserida!")
        shutil.copy2(bak, INDEX)
        sys.exit(1)

    # 7. Salva
    with open(INDEX, 'w', encoding='utf-8') as f:
        f.write(html)

    tamanho_depois = len(html)
    log(f"index.html salvo: {tamanho_depois:,} bytes (delta: {tamanho_depois - tamanho_antes:+,})")

    # 8. Confirmação final
    log("✅ Injeção cirúrgica concluída com sucesso!")
    log("   Abas verificadas:")
    for aba in abas_criticas + ['agenda-futura']:
        status = "✅" if f'id="{aba}"' in html else "❌"
        log(f"     {status} {aba}")


if __name__ == '__main__':
    main()
