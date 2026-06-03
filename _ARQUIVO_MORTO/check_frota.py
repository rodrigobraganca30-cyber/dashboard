import openpyxl
wb = openpyxl.load_workbook(r'C:\Users\SVOBODA\Desktop\DASHBOARD\frota_temp.xlsx', read_only=True, data_only=True)

# COMBUSTIVEL $ - analisar melhor a estrutura
ws = wb['COMBUSTÍVEL $']
rows = list(ws.iter_rows(values_only=True))
print('=== COMBUSTÍVEL $ - Detalhada ===')
# Row 8 parece ser header dos técnicos
for i in range(7, min(30, len(rows))):
    r = rows[i]
    nome = str(r[0])[:25] if r[0] else '-'
    locadora = str(r[1])[:15] if r[1] else '-'
    shell = str(r[2])[:15] if r[2] else '-'
    valecard = str(r[3])[:15] if r[3] else '-'
    soma = str(r[4])[:15] if r[4] else '-'
    litros = str(r[5])[:15] if r[5] else '-'
    km = str(r[6])[:15] if r[6] else '-'
    print(f'  {i+1}: {nome} | loc={locadora} | shell={shell} | vale={valecard} | soma={soma} | litros={litros} | km={km}')

# Verificar coluna H em diante (parece ter tabela pivot de valores)
print('\n=== Coluna H+ (Pivot) ===')
for i in range(3, min(15, len(rows))):
    r = rows[i]
    extras = [str(v)[:20] if v else '-' for v in r[7:12]]
    print(f'  Row {i+1}: {extras}')

# DESLOCAMENTO por mes
print('\n=== DESLOCAMENTO por Mês ===')
ws2 = wb['DESLOCAMENTO']
rows2 = list(ws2.iter_rows(values_only=True))
from collections import defaultdict
por_mes = defaultdict(lambda: {'km': 0, 'val': 0, 'count': 0})
for r in rows2[1:]:
    mes = str(r[0])[:7] if r[0] else '-'
    try: km = float(r[3])
    except: km = 0
    try: val = float(r[4])
    except: val = 0
    por_mes[mes]['km'] += km
    por_mes[mes]['val'] += val
    por_mes[mes]['count'] += 1
for m in sorted(por_mes.keys()):
    v = por_mes[m]
    print(f'  {m}: {v["count"]} viagens, {v["km"]:.0f} km, R$ {v["val"]:.2f}')

wb.close()
