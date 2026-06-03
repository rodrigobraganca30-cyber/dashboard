"""
injetar_os_data.py
------------------
Injeta APENAS os dados da planilha ATIVIDADES DAS O.S.xlsx
diretamente no index.html do servidor, sem regenerar o dashboard inteiro.
"""
import openpyxl, json, unicodedata, re, os, sys
from datetime import datetime
from collections import Counter

PASTA = os.path.dirname(os.path.abspath(__file__))
caminho_os = os.path.join(PASTA, "ATIVIDADES DAS O.S.xlsx")

if not os.path.exists(caminho_os):
    print(f"[ERRO] {caminho_os} não encontrado!")
    sys.exit(1)

print(f"[1/4] Lendo planilha: {os.path.basename(caminho_os)}")

def norm_str(v):
    if not v: return ""
    return str(v).strip().title()

def gv(row, i):
    return row[i] if i < len(row) else None

def _norm_hdr(h):
    if not h: return ""
    s = unicodedata.normalize('NFD', str(h).strip())
    return "".join(c for c in s if unicodedata.category(c) != 'Mn')

def _norm_status(s):
    n = unicodedata.normalize('NFD', str(s).lower())
    return "".join(c for c in n if unicodedata.category(c) != 'Mn')

# Carregar planilha
wb = openpyxl.load_workbook(caminho_os, read_only=True, data_only=True)
aba = next((n for n in ["Plan1", "BASE_DADOS", "Base"] if n in wb.sheetnames), wb.sheetnames[0])
ws = wb[aba]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]
idx = {_norm_hdr(h): i for i, h in enumerate(headers) if h}

CR = idx.get("Recurso", 1)
CD = idx.get("Data", 2)
CS = idx.get("Status da Atividade", 3)
CC = idx.get("Cidade", 6)
CT2 = idx.get("Tipo de Atividade_2", idx.get("Tipo de Atividade 2", idx.get("Tipo de Atividade", 24)))
if CT2 == idx.get("Tipo de Atividade") and CT2 is not None:
    if len(headers) > 24 and str(headers[24]).startswith("Tipo"):
        CT2 = 24
CMOT = idx.get("Motivo de Encerramento das atividades", 38)
CDUR = idx.get("Duracao", 21)
CDESL = idx.get("Tempo de Deslocamento", 22)
COS_NUM = idx.get("Atividade", idx.get("Numero da Atividade", idx.get("N da Atividade", 4)))

tags_tec, tags_city, tags_tipo, tags_dia = {}, {}, {}, {}
motivos_canc = Counter()
status_counts = {"Concluída": 0, "Cancelada": 0, "Não Concluída": 0, "Suspensa": 0, "Pendente": 0}
raw_os_data = []
tempos_duracao = []
tempos_deslocamento = []
sla_estourado = 0
total_os = 0

for row in rows[1:]:
    status_raw = gv(row, CS)
    if not status_raw: continue
    status_norm = _norm_status(status_raw)
    cat = "Pendente"
    if "nao conclu" in status_norm:   cat = "Não Concluída"
    elif "conclu" in status_norm:     cat = "Concluída"
    elif "cancel" in status_norm:     cat = "Cancelada"
    elif "suspen" in status_norm:     cat = "Suspensa"
    elif "iniciado" in status_norm or "em rota" in status_norm: cat = "Pendente"
    
    total_os += 1
    status_counts[cat] += 1
    
    recurso = norm_str(gv(row, CR))
    cidade = norm_str(gv(row, CC))
    tipo = gv(row, CT2)
    motivo = gv(row, CMOT)
    data_raw = row[CD] if len(row) > CD else None
    
    data_key = "--"
    if data_raw:
        try:
            dt_obj = data_raw if not isinstance(data_raw, str) else datetime.strptime(str(data_raw)[:10], "%Y-%m-%d")
            data_key = dt_obj.strftime("%d/%m")
        except: pass

    dur_raw = gv(row, CDUR)
    desl_raw = gv(row, CDESL)
    dur_min = 0
    desl_min = 0
    if dur_raw:
        try:
            dur_min = int(str(dur_raw).split(':')[0]) * 60 + int(str(dur_raw).split(':')[1]) if ':' in str(dur_raw) else int(float(str(dur_raw)))
            if dur_min > 0: tempos_duracao.append(dur_min)
        except: pass
    if desl_raw:
        try:
            desl_min = int(str(desl_raw).split(':')[0]) * 60 + int(str(desl_raw).split(':')[1]) if ':' in str(desl_raw) else int(float(str(desl_raw)))
            if desl_min > 0: tempos_deslocamento.append(desl_min)
        except: pass
    
    sla_flag = 0
    os_num = gv(row, COS_NUM)
    if cat == "Concluída":
        try:
            abertura = row[15]
            data_ativ = row[CD]
            hora_fim = row[17]
            if abertura and data_ativ and hora_fim:
                from datetime import timedelta as td
                dt_fim = datetime.combine(data_ativ.date() if hasattr(data_ativ, 'date') else data_ativ, hora_fim)
                diff_h = (dt_fim - abertura).total_seconds() / 3600
                if diff_h < 0:
                    dt_fim = datetime.combine((data_ativ.date() if hasattr(data_ativ, 'date') else data_ativ) + td(days=1), hora_fim)
                    diff_h = (dt_fim - abertura).total_seconds() / 3600
                if diff_h > 24:
                    sla_estourado += 1
                    sla_flag = 1
        except: pass

    raw_os_data.append({"tec": recurso, "city": cidade, "status": cat, "type": tipo, "date": data_key, "motivo": motivo, "dur": dur_min, "desl": desl_min, "sla": sla_flag, "os": os_num})
    if cat in ["Cancelada", "Não Concluída"] and motivo: motivos_canc[motivo] += 1

    for d, k in [(tags_tec, recurso), (tags_city, cidade), (tags_tipo, tipo)]:
        if k not in d: d[k] = {"key": k, "total": 0, "concluido": 0, "nao_concluido": 0, "cancelado": 0}
        d[k]["total"] += 1
        if cat == "Concluída": d[k]["concluido"] += 1
        elif cat == "Não Concluída": d[k]["nao_concluido"] += 1
        elif cat == "Cancelada": d[k]["cancelado"] += 1

    if data_key != "--":
        if data_key not in tags_dia: tags_dia[data_key] = {"data": data_key, "total": 0, "concluido": 0}
        tags_dia[data_key]["total"] += 1
        if cat == "Concluída": tags_dia[data_key]["concluido"] += 1

def calc_taxa(d):
    d["taxa"] = round((d["concluido"] / d["total"] * 100), 1) if d["total"] > 0 else 0
    if "key" in d:
        if d["key"] in tags_tec: d["Recurso"] = d["key"]
        elif d["key"] in tags_city: d["Cidade"] = d["key"]
        elif d["key"] in tags_tipo: d["tipo"] = d["key"]
    return d

list_tec = sorted([calc_taxa(v) for v in tags_tec.values()], key=lambda x: x["taxa"], reverse=True)
list_city = sorted([calc_taxa(v) for v in tags_city.values()], key=lambda x: x["total"], reverse=True)
list_tipo = sorted([calc_taxa(v) for v in tags_tipo.values()], key=lambda x: x["total"], reverse=True)
list_diag = sorted([calc_taxa(v) for v in tags_dia.values()], key=lambda x: x["data"])

datas_unicas = sorted(list(set(d["date"] for d in raw_os_data if d["date"] != "--")))

print(f"[2/4] Processado: {total_os} OS, {len(datas_unicas)} datas ({datas_unicas[0]} a {datas_unicas[-1]})")
print(f"       Status: C={status_counts['Concluída']} NC={status_counts['Não Concluída']} CA={status_counts['Cancelada']} SU={status_counts['Suspensa']} PE={status_counts['Pendente']}")

# Gerar JSON de dados para injeção
filtros_listas = {
    "periodos": ["Hoje", "Últimos 3 dias", "Últimos 5 dias", "Últimos 7 dias", "Últimos 10 dias"],
    "allDates": datas_unicas,
    "citys": sorted(list(set(d["city"] for d in raw_os_data))),
    "tecs": sorted(list(set(d["tec"] for d in raw_os_data))),
    "types": sorted(list(set(d["type"] for d in raw_os_data if d["type"])))
}

# Salvar JSON temporário para upload
data_json = {
    "tecData": list_tec,
    "cityData": list_city,
    "tipoData": list_tipo,
    "diaData": list_diag,
    "rawOSData": raw_os_data,
    "filterLists": filtros_listas,
    "status_counts": status_counts,
    "total_os": total_os,
    "sla_estourado": sla_estourado,
    "tempo_medio": round(sum(tempos_duracao) / len(tempos_duracao)) if tempos_duracao else 0,
    "desl_medio": round(sum(tempos_deslocamento) / len(tempos_deslocamento)) if tempos_deslocamento else 0,
    "periodo": f"{datas_unicas[0]} a {datas_unicas[-1]}" if datas_unicas else "sem dados",
    "motivos": dict(motivos_canc.most_common(10))
}

json_file = os.path.join(PASTA, "_os_inject.json")
with open(json_file, "w", encoding="utf-8") as f:
    json.dump(data_json, f, ensure_ascii=False)
print(f"[3/4] JSON gerado: {os.path.getsize(json_file):,} bytes")

# Upload JSON e executar injeção no servidor
print("[4/4] Enviando para servidor e injetando...")
