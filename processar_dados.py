"""
processar_dados.py — FASE 1: Gera JSONs a partir das planilhas Excel
NÃO modifica o HTML, NÃO substitui nada.
Apenas cria arquivos .json na pasta data/ para uso futuro.

Pode ser chamado APÓS o gerar_dashboard_v2.py ou independentemente.
"""
import openpyxl, os, json, sys, unicodedata, glob
from datetime import datetime
from collections import Counter

sys.stdout.reconfigure(encoding='utf-8', errors='replace')

PASTA = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(PASTA, "data")
os.makedirs(DATA_DIR, exist_ok=True)

def gv(row, col):
    try: return str(row[col]).strip() if row[col] not in (None, "", "None") else ""
    except: return ""

def norm_str(s):
    return str(s).strip().title() if s else ""

def _norm_hdr(h):
    if not h: return ""
    s = unicodedata.normalize('NFD', str(h).strip())
    return "".join(c for c in s if unicodedata.category(c) != 'Mn')

def _norm_status(s):
    n = unicodedata.normalize('NFD', str(s).lower())
    return "".join(c for c in n if unicodedata.category(c) != 'Mn')

# ═══════════════════════════════════════
# 1. ATIVIDADES (O.S.)
# ═══════════════════════════════════════
caminho_os = os.path.join(PASTA, "ATIVIDADES DAS O.S.xlsx")
# Fallback: no servidor VPS o arquivo fica em /opt/painel_robo/data/csv/
if not os.path.exists(caminho_os):
    caminho_os = "/opt/painel_robo/data/csv/ATIVIDADES DAS O.S.xlsx"
if not os.path.exists(caminho_os):
    print(f"[ERRO] Planilha nao encontrada em nenhum caminho")
    sys.exit(1)

print("[1/5] Processando ATIVIDADES DAS O.S.xlsx...")
wb = openpyxl.load_workbook(caminho_os, read_only=True, data_only=True)
aba = next((n for n in ["Plan1", "BASE_DADOS", "Base"] if n in wb.sheetnames), wb.sheetnames[0])
ws = wb[aba]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

idx = {_norm_hdr(h): i for i, h in enumerate(headers) if h}
CR = idx.get("Recurso", 1)
CD = idx.get("Data", 2)
CS = idx.get("Status da Atividade", 3)
CC = idx.get("Cidade", 6)
CT2 = idx.get("Tipo de Atividade_2", idx.get("Tipo de Atividade 2", idx.get("Tipo de Atividade", 24)))
if CT2 == idx.get("Tipo de Atividade") and CT2 is not None:
    if len(headers) > 24 and str(headers[24]).startswith("Tipo"):
        CT2 = 24
CMOT = idx.get("Motivo de Encerramento das atividades", 38)
CDUR = idx.get("Duracao", 21)
CDESL = idx.get("Tempo de Deslocamento", 22)
COS_NUM = idx.get("Atividade", idx.get("Numero da Atividade", idx.get("N da Atividade", 4)))

raw_os_data = []
status_counts = {"Concluida": 0, "Cancelada": 0, "Nao Concluida": 0, "Suspensa": 0, "Pendente": 0}
motivos_canc = Counter()
tags_tec, tags_city, tags_tipo, tags_dia = {}, {}, {}, {}
tempos_duracao = []
tempos_deslocamento = []
sla_estourado = 0
total_os = 0

for row in rows[1:]:
    status_raw = gv(row, CS)
    if not status_raw: continue
    status_norm = _norm_status(status_raw)
    cat = "Pendente"
    if "nao conclu" in status_norm:   cat = "Nao Concluida"
    elif "conclu" in status_norm:     cat = "Concluida"
    elif "cancel" in status_norm:     cat = "Cancelada"
    elif "suspen" in status_norm:     cat = "Suspensa"
    elif "iniciado" in status_norm or "em rota" in status_norm: cat = "Pendente"

    total_os += 1
    status_counts[cat] = status_counts.get(cat, 0) + 1

    recurso = norm_str(gv(row, CR))
    cidade = norm_str(gv(row, CC))
    tipo = gv(row, CT2)
    motivo = gv(row, CMOT)
    data_raw = row[CD] if len(row) > CD else None

    data_key = "--"
    if data_raw:
        try:
            dt_obj = data_raw if not isinstance(data_raw, str) else datetime.strptime(str(data_raw)[:10], "%Y-%m-%d")
            data_key = dt_obj.strftime("%d/%m")
        except: pass

    dur_min = 0
    desl_min = 0
    dur_raw = gv(row, CDUR)
    desl_raw = gv(row, CDESL)
    if dur_raw:
        try:
            dur_min = int(str(dur_raw).split(':')[0]) * 60 + int(str(dur_raw).split(':')[1]) if ':' in str(dur_raw) else int(float(str(dur_raw)))
            if dur_min > 0: tempos_duracao.append(dur_min)
        except: pass
    if desl_raw:
        try:
            desl_min = int(str(desl_raw).split(':')[0]) * 60 + int(str(desl_raw).split(':')[1]) if ':' in str(desl_raw) else int(float(str(desl_raw)))
            if desl_min > 0: tempos_deslocamento.append(desl_min)
        except: pass

    sla_flag = 0
    os_num = gv(row, COS_NUM)
    if cat == "Concluida":
        try:
            abertura = row[15]
            data_ativ = row[CD]
            hora_fim = row[17]
            if abertura and data_ativ and hora_fim:
                from datetime import timedelta as td
                dt_fim = datetime.combine(data_ativ.date() if hasattr(data_ativ, 'date') else data_ativ, hora_fim)
                diff_h = (dt_fim - abertura).total_seconds() / 3600
                if diff_h < 0:
                    dt_fim = datetime.combine((data_ativ.date() if hasattr(data_ativ, 'date') else data_ativ) + td(days=1), hora_fim)
                    diff_h = (dt_fim - abertura).total_seconds() / 3600
                if diff_h > 24:
                    sla_estourado += 1
                    sla_flag = 1
        except: pass

    raw_os_data.append({"tec": recurso, "city": cidade, "status": cat, "type": tipo, "date": data_key, "motivo": motivo, "dur": dur_min, "desl": desl_min, "sla": sla_flag, "os": os_num})
    if cat in ["Cancelada", "Nao Concluida"] and motivo: motivos_canc[motivo] += 1

    for d, k in [(tags_tec, recurso), (tags_city, cidade), (tags_tipo, tipo)]:
        if k not in d: d[k] = {"key": k, "total": 0, "concluido": 0, "nao_concluido": 0, "cancelado": 0}
        d[k]["total"] += 1
        if cat == "Concluida": d[k]["concluido"] += 1
        elif cat == "Nao Concluida": d[k]["nao_concluido"] += 1
        elif cat == "Cancelada": d[k]["cancelado"] += 1

    if data_key != "--":
        if data_key not in tags_dia: tags_dia[data_key] = {"data": data_key, "total": 0, "concluido": 0}
        tags_dia[data_key]["total"] += 1
        if cat == "Concluida": tags_dia[data_key]["concluido"] += 1

wb.close()

def calc_taxa(d):
    d["taxa"] = round((d["concluido"] / d["total"] * 100), 1) if d["total"] > 0 else 0
    return d

list_tec = sorted([calc_taxa(v) for v in tags_tec.values()], key=lambda x: x["taxa"], reverse=True)
list_city = sorted([calc_taxa(v) for v in tags_city.values()], key=lambda x: x["total"], reverse=True)
list_tipo = sorted([calc_taxa(v) for v in tags_tipo.values()], key=lambda x: x["total"], reverse=True)
list_diag = sorted([calc_taxa(v) for v in tags_dia.values()], key=lambda x: x["data"])
datas_unicas = sorted(set(d["date"] for d in raw_os_data if d["date"] != "--"))

# Recordes
v_tec_taxa = list_tec[0] if list_tec else {"key": "-", "taxa": 0}
v_tec_prod = sorted(list_tec, key=lambda x: x["total"], reverse=True)[0] if list_tec else {"key": "-", "total": 0}
v_tec_bad = list_tec[-1] if list_tec else {"key": "-", "taxa": 0}
v_cit_taxa = sorted(list_city, key=lambda x: x["taxa"], reverse=True)[0] if list_city else {"key": "-", "taxa": 0}
v_cit_vol = list_city[0] if list_city else {"key": "-", "total": 0}
v_cit_bad = sorted(list_city, key=lambda x: x["taxa"])[0] if list_city else {"key": "-", "taxa": 0}
v_pico = sorted(list_diag, key=lambda x: x["total"], reverse=True)[0] if list_diag else {"data": "-", "total": 0}
v_melhor = sorted(list_diag, key=lambda x: x["taxa"], reverse=True)[0] if list_diag else {"data": "-", "taxa": 0}
v_pior = sorted(list_diag, key=lambda x: x["taxa"])[0] if list_diag else {"data": "-", "taxa": 0}
v_media = round(total_os / len(list_diag), 1) if list_diag else 0

# Salvar os.json
with open(os.path.join(DATA_DIR, "os.json"), "w", encoding="utf-8") as f:
    json.dump(raw_os_data, f, ensure_ascii=False)
print(f"    os.json: {len(raw_os_data)} registros")

# Salvar filtros.json
filtros = {
    "periodos": ["Hoje", "Ultimos 3 dias", "Ultimos 5 dias", "Ultimos 7 dias", "Ultimos 10 dias"],
    "allDates": datas_unicas,
    "citys": sorted(list(set(d["city"] for d in raw_os_data))),
    "tecs": sorted(list(set(d["tec"] for d in raw_os_data))),
    "types": sorted(list(set(d["type"] for d in raw_os_data)))
}
with open(os.path.join(DATA_DIR, "filtros.json"), "w", encoding="utf-8") as f:
    json.dump(filtros, f, ensure_ascii=False)
print(f"    filtros.json: {len(filtros['citys'])} cidades, {len(filtros['tecs'])} tecnicos")

# Salvar kpis.json
kpis = {
    "total_os": total_os,
    "status_counts": status_counts,
    "motivos_canc": dict(motivos_canc.most_common(20)),
    "media_duracao": round(sum(tempos_duracao) / len(tempos_duracao), 1) if tempos_duracao else 0,
    "media_deslocamento": round(sum(tempos_deslocamento) / len(tempos_deslocamento), 1) if tempos_deslocamento else 0,
    "sla_estourado": sla_estourado,
    "recordes": {
        "tec_melhor_taxa": {"nome": v_tec_taxa.get("key", "-"), "taxa": v_tec_taxa.get("taxa", 0)},
        "tec_mais_prod": {"nome": v_tec_prod.get("key", "-"), "total": v_tec_prod.get("total", 0)},
        "tec_pior_taxa": {"nome": v_tec_bad.get("key", "-"), "taxa": v_tec_bad.get("taxa", 0)},
        "cit_melhor_taxa": {"nome": v_cit_taxa.get("key", "-"), "taxa": v_cit_taxa.get("taxa", 0)},
        "cit_mais_vol": {"nome": v_cit_vol.get("key", "-"), "total": v_cit_vol.get("total", 0)},
        "cit_pior_taxa": {"nome": v_cit_bad.get("key", "-"), "taxa": v_cit_bad.get("taxa", 0)},
        "pico_dia": {"data": v_pico.get("data", "-"), "total": v_pico.get("total", 0)},
        "melhor_dia": {"data": v_melhor.get("data", "-"), "taxa": v_melhor.get("taxa", 0)},
        "pior_dia": {"data": v_pior.get("data", "-"), "taxa": v_pior.get("taxa", 0)},
        "media_diaria": v_media
    },
    "gerado_em": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
}
with open(os.path.join(DATA_DIR, "kpis.json"), "w", encoding="utf-8") as f:
    json.dump(kpis, f, ensure_ascii=False, indent=2)
print(f"    kpis.json: total={total_os}")

# ═══════════════════════════════════════
# 2. FROTA
# ═══════════════════════════════════════
print("\n[2/5] Processando frota...")
FROTA_FILE = os.path.join(PASTA, "frota_temp.xlsx")
FROTA_SHEET_ID = "1CjWOfMTQQ1aP3jeF-Wf3gxIa5DT0qMRwjijD6nXXtQg"
FROTA_URL = f"https://docs.google.com/spreadsheets/d/{FROTA_SHEET_ID}/export?format=xlsx"

# Download automático da planilha de frota
try:
    import urllib.request, shutil
    frota_temp_dl = os.path.join(PASTA, "frota_download.xlsx")
    print("    Baixando frota do Google Sheets...")
    urllib.request.urlretrieve(FROTA_URL, frota_temp_dl)
    shutil.move(frota_temp_dl, FROTA_FILE)
    print("    [OK] frota_temp.xlsx atualizado!")
except Exception as e:
    print(f"    [AVISO] Nao foi possivel baixar frota: {e}")
    print("    [INFO] Usando arquivo local existente")

frota_data = {"veiculos": [], "oficina": [], "revisao": [], "combustivel": [], "deslocamento_tec": [], "deslocamento_dest": [], "kpis": {}}

if os.path.exists(FROTA_FILE):
    wb_frota = openpyxl.load_workbook(FROTA_FILE, read_only=True, data_only=True)

    # Veiculos
    if "PLACA X TECNICOS" in [_norm_hdr(n) for n in wb_frota.sheetnames]:
        ws_name = next(n for n in wb_frota.sheetnames if _norm_hdr(n) == "PLACA X TECNICOS")
        ws_f = wb_frota[ws_name]
        for r in list(ws_f.iter_rows(values_only=True))[1:]:
            placa = str(r[0]).strip() if r[0] else ""
            func = str(r[1]).strip() if r[1] else ""
            locadora = str(r[3]).strip() if r[3] else ""
            status = str(r[4]).strip() if r[4] else ""
            if placa and placa != "None":
                frota_data["veiculos"].append({"placa": placa, "tecnico": func, "locadora": locadora, "status": status})

    # Oficina
    oficina_nome = next((n for n in wb_frota.sheetnames if "OFICINA" in n.upper()), None)
    if oficina_nome:
        ws_o = wb_frota[oficina_nome]
        for r in list(ws_o.iter_rows(values_only=True))[1:]:
            status_o = str(r[6]).strip() if r[6] else ""
            if status_o and status_o.upper() not in ["N/A", "None", "PRONTO", "RETIRADO", ""] and "OFICINA" in status_o.upper():
                placa = str(r[0]).strip() if r[0] else ""
                motorista = str(r[1]).strip() if r[1] else ""
                entrada = str(r[3])[:10] if r[3] else "-"
                previsao = str(r[4])[:10] if r[4] else "-"
                cidade = str(r[7]).strip() if r[7] else ""
                obs = str(r[8]).strip() if r[8] else ""
                dias = 0
                try:
                    dt_entrada = r[3] if hasattr(r[3], 'date') else datetime.strptime(str(r[3])[:10], '%Y-%m-%d')
                    saida_raw = r[5]
                    if saida_raw and str(saida_raw).strip() not in ['', 'None', '-']:
                        dt_saida = saida_raw if hasattr(saida_raw, 'date') else datetime.strptime(str(saida_raw)[:10], '%Y-%m-%d')
                    else:
                        dt_saida = datetime.now()
                    dias = max(0, (dt_saida - dt_entrada).days)
                except: pass
                if placa and placa != "None":
                    frota_data["oficina"].append({"placa": placa, "motorista": motorista, "entrada": entrada, "previsao": previsao, "status": status_o, "cidade": cidade, "obs": obs, "dias": dias})

    # Revisao
    revisao_nome = next((n for n in wb_frota.sheetnames if "REVIS" in n.upper()), None)
    if revisao_nome:
        ws_r = wb_frota[revisao_nome]
        for r in list(ws_r.iter_rows(values_only=True))[1:]:
            rev_status = str(r[5]).strip() if r[5] else ""
            if rev_status in ["REVISAO", "AGENDAR", "REVISAR", "REVISÃO"]:
                tec = str(r[0]).strip() if r[0] else ""
                placa = str(r[1]).strip() if r[1] else ""
                if placa.upper() == "PLACA": continue
                locadora = str(r[2]).strip() if r[2] else ""
                km = str(r[3]).split(".")[0] if r[3] else "-"
                data_rev = str(r[4])[:10] if r[4] else "-"
                if placa and placa != "None":
                    frota_data["revisao"].append({"tecnico": tec, "placa": placa, "locadora": locadora, "km": km, "data": data_rev, "status": rev_status})

    # Combustivel
    comb_nome = next((n for n in wb_frota.sheetnames if "COMBUST" in n.upper()), None)
    total_gasto_comb = total_litros_comb = total_km_comb = 0
    if comb_nome:
        ws_c = wb_frota[comb_nome]
        for r in list(ws_c.iter_rows(values_only=True)):
            nome = str(r[0]).strip() if r[0] else ""
            if not nome or nome in ["-", "None", "l"]: continue
            if "equipe" in nome.lower() or "BH" in nome or "=" in nome or "Home" in nome or "ALMOX" in nome.upper(): continue
            try:
                soma = float(str(r[4]).replace("R$","").replace(",",".").strip()) if r[4] and str(r[4]).strip() not in ["-", "None", "SOMA", "SOMA "] else 0
                litros = float(r[5]) if r[5] and str(r[5]).strip() not in ["-", "None", "LITROS", "LITROS "] else 0
                km = float(r[6]) if r[6] and str(r[6]).strip() not in ["-", "None", "KM", "KM "] else 0
            except: soma = litros = km = 0
            if soma > 0:
                frota_data["combustivel"].append({"tecnico": nome, "gasto": round(soma, 2), "litros": round(litros, 1), "km": round(km, 0)})
                total_gasto_comb += soma
                total_litros_comb += litros
                total_km_comb += km
    frota_data["combustivel"].sort(key=lambda x: x["gasto"], reverse=True)

    # Deslocamento
    desl_nome = next((n for n in wb_frota.sheetnames if "DESLOCAMENTO" in n.upper()), None)
    desl_por_tec = {}
    desl_por_dest = {}
    total_km_desl = total_val_desl = 0
    if desl_nome:
        ws_d = wb_frota[desl_nome]
        for r in list(ws_d.iter_rows(values_only=True))[1:]:
            tec_d = str(r[1]).strip() if r[1] else ""
            dest = str(r[2]).strip() if r[2] else ""
            try: km_d = float(r[3])
            except: km_d = 0
            try: val_d = float(r[4])
            except: val_d = 0
            if tec_d and tec_d != "None":
                if tec_d not in desl_por_tec: desl_por_tec[tec_d] = {"viagens": 0, "km": 0, "valor": 0}
                desl_por_tec[tec_d]["viagens"] += 1
                desl_por_tec[tec_d]["km"] += km_d
                desl_por_tec[tec_d]["valor"] += val_d
            if dest and dest != "None":
                if dest not in desl_por_dest: desl_por_dest[dest] = {"viagens": 0, "km": 0}
                desl_por_dest[dest]["viagens"] += 1
                desl_por_dest[dest]["km"] += km_d
            total_km_desl += km_d
            total_val_desl += val_d

    for t, v in sorted(desl_por_tec.items(), key=lambda x: x[1]["km"], reverse=True):
        frota_data["deslocamento_tec"].append({"tecnico": t, "viagens": v["viagens"], "km": round(v["km"], 0), "valor": round(v["valor"], 2)})
    for d, v in sorted(desl_por_dest.items(), key=lambda x: x[1]["viagens"], reverse=True):
        frota_data["deslocamento_dest"].append({"destino": d, "viagens": v["viagens"], "km": round(v["km"], 0)})

    veic_total = len(frota_data["veiculos"])
    veic_ativos = sum(1 for v in frota_data["veiculos"] if "ATIVIDADE" in v["status"].upper())
    frota_data["kpis"] = {
        "total": veic_total,
        "ativos": veic_ativos,
        "oficina": len(frota_data["oficina"]),
        "revisao": len(frota_data["revisao"]),
        "combustivel_total": round(total_gasto_comb, 2),
        "litros_total": round(total_litros_comb, 1),
        "km_total": round(total_km_comb, 0),
        "deslocamento_km": round(total_km_desl, 0),
        "deslocamento_valor": round(total_val_desl, 2),
    }
    wb_frota.close()
else:
    print("    [AVISO] frota_temp.xlsx nao encontrado")

with open(os.path.join(DATA_DIR, "frota.json"), "w", encoding="utf-8") as f:
    json.dump(frota_data, f, ensure_ascii=False)
print(f"    frota.json: {len(frota_data['veiculos'])} veiculos, {len(frota_data['oficina'])} em oficina")

# ═══════════════════════════════════════
# 3. INDICADORES GIGA
# ═══════════════════════════════════════
print("\n[3/5] Processando Indicadores Giga...")
GIGA_FILE = os.path.join(PASTA, "giga_temp.xlsx")
GIGA_SHEET_ID = "1cWX9jvJ1WaAFDkWi9aulZ675tMxEkZrmqdnZSYwcAWc"
GIGA_URL = f"https://docs.google.com/spreadsheets/d/{GIGA_SHEET_ID}/export?format=xlsx"

try:
    giga_dl = os.path.join(PASTA, "giga_download.xlsx")
    print("    Baixando Giga do Google Sheets...")
    urllib.request.urlretrieve(GIGA_URL, giga_dl)
    shutil.move(giga_dl, GIGA_FILE)
    print("    [OK] giga_temp.xlsx atualizado!")
except Exception as e:
    print(f"    [AVISO] Nao foi possivel baixar Giga: {e}")

giga_data = []

if os.path.exists(GIGA_FILE):
    wb_giga = openpyxl.load_workbook(GIGA_FILE, read_only=True, data_only=True)
    for sn in wb_giga.sheetnames:
        ws = wb_giga[sn]
        rows_g = list(ws.iter_rows(values_only=True))
        if len(rows_g) > 1:
            hdr = rows_g[0]
            for r in rows_g[1:]:
                entry = {}
                for ci, h in enumerate(hdr):
                    if h and ci < len(r):
                        val = r[ci]
                        if val is not None:
                            entry[str(h).strip()] = str(val) if not isinstance(val, (int, float)) else val
                if entry:
                    entry["_sheet"] = sn
                    giga_data.append(entry)
    wb_giga.close()
else:
    print("    [AVISO] giga_temp.xlsx nao encontrado")

with open(os.path.join(DATA_DIR, "giga.json"), "w", encoding="utf-8") as f:
    json.dump(giga_data, f, ensure_ascii=False)
print(f"    giga.json: {len(giga_data)} registros")

print("\n[4/5] Processando Premios (Indicador Giga)...")
PREMIO_FILE = os.path.join(PASTA, "premio_temp.xlsx")
GIGA2_SHEET_ID = "121xl-QAkm8MNYM6hyQk2ueV2uBIybP_37O22zlj9HYg"
GIGA2_URL = f"https://docs.google.com/spreadsheets/d/{GIGA2_SHEET_ID}/export?format=xlsx"

try:
    premio_dl = os.path.join(PASTA, "premio_download.xlsx")
    print("    Baixando Premio do Google Sheets...")
    urllib.request.urlretrieve(GIGA2_URL, premio_dl)
    shutil.move(premio_dl, PREMIO_FILE)
    print("    [OK] premio_temp.xlsx atualizado!")
except Exception as e:
    print(f"    [AVISO] Nao foi possivel baixar Premio: {e}")

# Gerar giga2.json no formato estruturado (mesmo que _GIGA2 do gerar_dashboard_v2.py)
import re as _re2
_IND_INV = {"IFI", "IRR", "IFIME", "OUTLIER"}
def _is_inv(nome):
    n = nome.upper()
    return any(k in n for k in _IND_INV)

def _parse_meta(s):
    m = _re2.search(r'[\d,.]+', str(s).replace(',','.'))
    return float(m.group()) if m else 0.0

giga2_data = {}
if os.path.exists(PREMIO_FILE):
    _wb2 = openpyxl.load_workbook(PREMIO_FILE, read_only=True, data_only=True)
    if "PAINEL" in _wb2.sheetnames:
        _ws2 = _wb2["PAINEL"]
        _rows2 = list(_ws2.iter_rows(values_only=True))
        _mes2 = None
        _i2 = 0
        while _i2 < len(_rows2):
            _r2 = _rows2[_i2]
            _v0 = str(_r2[0]).strip() if _r2[0] else ""
            if not _v0:
                _i2 += 1; continue
            # Linha de MES
            if not _v0.startswith("C.") and _v0 not in ["INDICADORES", "TOTAL"] \
               and not any(_k in _v0.upper() for _k in ["EFICI","AGING","IFI","PRAZO","OUTL","IRR"]):
                _mes2 = _v0
                if _mes2 not in giga2_data: giga2_data[_mes2] = {}
                _i2 += 1; continue
            # Linha de CENTRAL
            if _v0.startswith("C.") and _mes2:
                _central2 = _v0.split(",")[0].strip()
                _datas2 = []
                _col2 = 4
                while _col2 < len(_r2):
                    _dt2 = _r2[_col2]
                    if _dt2 is None: break
                    if hasattr(_dt2, 'day'): _datas2.append(f"{_dt2.day:02d}/{_dt2.month:02d}")
                    _col2 += 2
                _inds2 = []
                for _j2 in range(_i2 + 2, min(_i2 + 11, len(_rows2))):
                    _ri2 = _rows2[_j2]
                    _ni2 = str(_ri2[0]).strip() if _ri2[0] else ""
                    if not _ni2 or "TOTAL" in _ni2.upper(): break
                    try: _pm2 = int(float(_ri2[3])) if _ri2[3] else 0
                    except: _pm2 = 0
                    _dias2 = []
                    for _di2 in range(len(_datas2)):
                        _cn2 = 4 + _di2 * 2; _cp2 = _cn2 + 1
                        _nota2 = float(_ri2[_cn2]) if _cn2 < len(_ri2) and _ri2[_cn2] is not None else None
                        _peso2 = float(_ri2[_cp2]) if _cp2 < len(_ri2) and _ri2[_cp2] is not None else 0.0
                        _dias2.append({"n": round(_nota2 * 100, 1) if _nota2 is not None else None, "p": int(_peso2)})
                    _m100s = str(_ri2[1]).strip() if _ri2[1] else ""
                    _m80s = str(_ri2[2]).strip() if _ri2[2] else ""
                    _inds2.append({"nome": _ni2, "m100": _m100s, "m80": _m80s,
                                   "m100v": _parse_meta(_m100s), "m80v": _parse_meta(_m80s),
                                   "pmax": _pm2, "inv": _is_inv(_ni2), "dias": _dias2})
                _tidx2 = _i2 + 2 + len(_inds2)
                _tots2 = []
                if _tidx2 < len(_rows2):
                    _rt2 = _rows2[_tidx2]
                    for _di2 in range(len(_datas2)):
                        _ct2 = 4 + _di2 * 2
                        _vt2 = _rt2[_ct2] if _ct2 < len(_rt2) else None
                        _tots2.append(int(_vt2) if _vt2 is not None else 0)
                giga2_data[_mes2][_central2] = {"datas": _datas2, "inds": _inds2, "tots": _tots2}
                _i2 = _tidx2 + 1; continue
            _i2 += 1
        _wb2.close()
    print(f"    giga2.json: {list(giga2_data.keys())} meses, {sum(len(v) for v in giga2_data.values())} centrais")
else:
    print("    [AVISO] premio_temp.xlsx nao encontrado")

with open(os.path.join(DATA_DIR, "giga2.json"), "w", encoding="utf-8") as f:
    json.dump(giga2_data, f, ensure_ascii=False)
print(f"    giga2.json salvo")

# ═══════════════════════════════════════
# 5. RESUMO
# ═══════════════════════════════════════
print("\n[5/5] Resumo:")
for fname in ["os.json", "filtros.json", "kpis.json", "frota.json", "giga.json", "giga2.json"]:
    fpath = os.path.join(DATA_DIR, fname)
    size = os.path.getsize(fpath)
    print(f"    {fname}: {size:,} bytes ({size/1024:.1f} KB)")

total_size = sum(os.path.getsize(os.path.join(DATA_DIR, f)) for f in os.listdir(DATA_DIR) if f.endswith('.json'))
print(f"\n    TOTAL: {total_size:,} bytes ({total_size/1024:.1f} KB)")
print("\n[OK] Fase 1 concluida! JSONs gerados em:", DATA_DIR)
