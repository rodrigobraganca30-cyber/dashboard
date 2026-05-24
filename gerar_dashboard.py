import openpyxl,glob,os,json,re
from datetime import datetime

PASTA=r"C:\Users\SVOBODA\Desktop\DASHBOARD"
HTML_SRC=r"C:\Users\SVOBODA\Downloads\dashboard_svoboda.html"
HTML_OUT=os.path.join(PASTA,"dashboard_svoboda_atualizado.html")

arquivos=glob.glob(os.path.join(PASTA,"*.xlsx"))
caminho=max(arquivos,key=os.path.getmtime)
print("Lendo:",os.path.basename(caminho))

wb=openpyxl.load_workbook(caminho,read_only=True,data_only=True)
aba=next((n for n in ["Plan1","BASE_DADOS"] if n in wb.sheetnames),wb.sheetnames[0])
ws=wb[aba]
rows=list(ws.iter_rows(values_only=True))
headers=rows[0]
idx={str(h).strip():i for i,h in enumerate(headers) if h}

CR=idx.get("Recurso",1);CD=idx.get("Data",2);CS=idx.get("Status da Atividade",3)
CC=idx.get("Cidade",6);CT=idx.get("Tipo de Atividade_2",24)
CM=idx.get("Motivo de Encerramento das atividades",38)

tecs={};cids={};tips={};dias={};dmin=None;dmax=None;sts={}

def gv(r,c):
    try:
        v=r[c];return str(v).strip() if v is not None else ""
    except:return ""

for row in rows[1:]:
    s=gv(row,CS).lower();c=gv(row,CC).upper()
    t=gv(row,CT);r=gv(row,CR)
    d=row[CD] if len(row)>CD else None
    sts[s]=sts.get(s,0)+1

    for dic,key in [(cids,c),(tips,t),(tecs,r)]:
        if not key:continue
        if key not in dic:dic[key]={"t":0,"c":0,"n":0,"x":0}
        dic[key]["t"]+=1
        if "conclu" in s:dic[key]["c"]+=1
        elif "nao" in s or "não" in s:dic[key]["n"]+=1
        elif s=="cancelado":dic[key]["x"]+=1

    if d:
        try:
            dt=d if not isinstance(d,str) else datetime.strptime(str(d)[:10],"%Y-%m-%d")
            dk=dt.strftime("%d/%m")
            if dk not in dias:dias[dk]={"t":0,"c":0,"o":dt}
            dias[dk]["t"]+=1
            if "conclu" in s:dias[dk]["c"]+=1
            if dmin is None or dt<dmin:dmin=dt
            if dmax is None or dt>dmax:dmax=dt
        except:pass

def tx(c,t):return round(c/t*100,1) if t else 0

N=len(rows)-1
nc=sum(v for k,v in sts.items() if "conclu" in k)
nk=sts.get("cancelado",0)
nn=sum(v for k,v in sts.items() if "nao" in k or "não" in k)
ns=sts.get("suspenso",0);np_=sts.get("pendente",0);tg=tx(nc,N)

TL=sorted([{"Recurso":k,"total":v["t"],"concluido":v["c"],"nao_concluido":v["n"],"cancelado":v["x"],"taxa":tx(v["c"],v["t"])}
    for k,v in tecs.items() if v["t"]>=5],key=lambda x:-x["taxa"])
CL=sorted([{"Cidade":k,"total":v["t"],"concluido":v["c"],"taxa":tx(v["c"],v["t"])}
    for k,v in cids.items() if v["t"]>=10 and k.strip()],key=lambda x:-x["total"])
TiL=sorted([{"tipo":k,"total":v["t"],"concluido":v["c"],"nao_concluido":v["n"],"cancelado":v["x"],"taxa":tx(v["c"],v["t"])}
    for k,v in tips.items() if k.strip()],key=lambda x:-x["total"])
DL=[{"data":dk,"total":d["t"],"concluido":d["c"],"taxa":tx(d["c"],d["t"])}
    for dk,d in sorted(dias.items(),key=lambda x:x[1]["o"])]

periodo=(dmin.strftime("%d/%m/%Y")+" - "+dmax.strftime("%d/%m/%Y")) if dmin and dmax else ""
print("Agregado: "+str(N)+" OS | "+str(tg)+"% | "+str(len(TL))+" tecs | "+periodo)

with open(HTML_SRC,"r",encoding="utf-8") as f:html=f.read()

out=re.sub(r"const tecData\s*=\s*\[.*?\];","const tecData = "+json.dumps(TL,ensure_ascii=False)+";",html,flags=re.DOTALL)
out=re.sub(r"const cityData\s*=\s*\[.*?\];","const cityData = "+json.dumps(CL,ensure_ascii=False)+";",out,flags=re.DOTALL)
out=re.sub(r"const tipoData\s*=\s*\[.*?\];","const tipoData = "+json.dumps(TiL,ensure_ascii=False)+";",out,flags=re.DOTALL)
out=re.sub(r"const diaData\s*=\s*\[.*?\];","const diaData = "+json.dumps(DL,ensure_ascii=False)+";",out,flags=re.DOTALL)

for old,new in [
    (">2.851<",">"+format(N,",")+  "<"),
    (">1.426<",">"+format(nc,",")+ "<"),
    ("taxa: 50,0%","taxa: "+str(tg)+"%"),
    (">925<",">"+str(nk)+"<"),
    ("32,4% do total",str(tx(nk,N))+"% do total"),
    (">259<",">"+str(nn)+"<"),
    ("9,1% do total",str(tx(nn,N))+"% do total"),
    (">141<",">"+str(ns)+"<"),
    ("4,9% do total",str(tx(ns,N))+"% do total"),
    (">100<",">"+str(np_)+"<"),
    ("3,5% do total",str(tx(np_,N))+"% do total"),
    (">50%<",">"+str(tg)+"%<"),
    (u"útimos 8 dias úteis","período completo"),
    (u"últimos 8 dias úteis","período completo"),
    ("Dados: 25/03 \u2013 01/04/2026","Dados: "+periodo),
]:
    if old in out:out=out.replace(old,new,1)

with open(HTML_OUT,"w",encoding="utf-8") as f:f.write(out)
print("GERADO: "+HTML_OUT)
