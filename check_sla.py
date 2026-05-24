import openpyxl, glob, os
from datetime import datetime, timedelta

pasta = r'C:\Users\SVOBODA\Desktop\DASHBOARD'
arqs = [f for f in glob.glob(os.path.join(pasta, '*.xlsx')) if not os.path.basename(f).startswith('~$') and 'frota' not in f]
wb = openpyxl.load_workbook(max(arqs, key=os.path.getmtime), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))

sla_por_tipo = {}
for row in rows[1:]:
    status = str(row[3]).lower() if row[3] else ''
    if 'conclu' not in status or 'nao' in status or 'não' in status:
        continue
    tipo = str(row[24]) if row[24] else 'Outros'
    if tipo not in sla_por_tipo:
        sla_por_tipo[tipo] = {'total': 0, 'sla': 0}
    sla_por_tipo[tipo]['total'] += 1
    try:
        abertura = row[15]
        data_ativ = row[2]
        hora_fim = row[17]
        if abertura and data_ativ and hora_fim:
            dt_fim = datetime.combine(data_ativ.date(), hora_fim)
            diff = (dt_fim - abertura).total_seconds() / 3600
            if diff < 0:
                dt_fim = datetime.combine(data_ativ.date() + timedelta(days=1), hora_fim)
                diff = (dt_fim - abertura).total_seconds() / 3600
            if diff > 24:
                sla_por_tipo[tipo]['sla'] += 1
    except:
        pass

print("=== SLA > 24h por Tipo de Atividade (apenas Concluídas) ===")
total_sla = 0
total_conc = 0
for t, v in sla_por_tipo.items():
    pct = round(v['sla'] / v['total'] * 100) if v['total'] > 0 else 0
    print(f"  {t}: {v['sla']}/{v['total']} ({pct}%)")
    total_sla += v['sla']
    total_conc += v['total']
pct_total = round(total_sla / total_conc * 100) if total_conc > 0 else 0
print(f"\n  TOTAL: {total_sla}/{total_conc} ({pct_total}%)")
